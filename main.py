# carts_management_flet.py
import flet as ft
import sqlite3
from datetime import datetime
import os
import shutil
from contextlib import contextmanager
import threading
import time
import json
from pathlib import Path
import base64
from dotenv import load_dotenv
import random

# تحميل المتغيرات البيئية
load_dotenv()

# محاولة استيراد مكتبة MEGA
try:
    from mega import Mega, MegaRequestException
    MEGA_AVAILABLE = True
except ImportError:
    MEGA_AVAILABLE = False

# محاولة استيراد PIL للتعامل مع الصور
try:
    from PIL import Image
    import io
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# محاولة استيراد مكتبات إنشاء ملفات PDF
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# محاولة استيراد openpyxl للتصدير إلى Excel
try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# إعدادات قاعدة البيانات
DB_NAME = 'carts_management.db'
DEFAULT_USER = 'سعود'
DEFAULT_PASSWORD = '123456'
APP_NAME = "نظام إدارة العربات اليدوية - الحرم المكي الشريف"

# إعدادات MEGA - من المتغيرات البيئية
MEGA_EMAIL = os.getenv('MEGA_EMAIL', '')
MEGA_PASSWORD = os.getenv('MEGA_PASSWORD', '')

# قائمة المستودعات الأساسية
WAREHOUSES = [
    {'name': 'المستودع الرئيسي', 'capacity': 5000, 'type': 'main', 'description': 'المستودع الرئيسي الكبير خارج المنطقة المركزية'},
    {'name': 'المستودع الخارجي', 'capacity': 1500, 'type': 'external', 'description': 'المستودع المركزي المتوسط الحجم'},
    {'name': 'مركز التشغيل الشمالي', 'capacity': 500, 'type': 'north', 'description': 'مركز التشغيل الشمالي'},
    {'name': 'مركز التشغيل الجنوبي', 'capacity': 500, 'type': 'south', 'description': 'مركز التشغيل الجنوبي'}
]

# حالات العربات
CART_STATUS = {
    'sound': 'سليمة',
    'needs_maintenance': 'تحتاج صيانة',
    'damaged': 'تالفة'
}

# حالات الصيانة
MAINTENANCE_STATUS = {
    'pending': 'بانتظار الصيانة',
    'in_progress': 'قيد التنفيذ',
    'completed': 'منجزة'
}

# الصلاحيات الافتراضية
DEFAULT_PERMISSIONS = {
    'can_view_dashboard': 1,
    'can_manage_carts': 1,
    'can_add_cart': 1,
    'can_edit_cart': 0,
    'can_delete_cart': 0,
    'can_move_cart': 1,
    'can_view_movements': 1,
    'can_manage_maintenance': 1,
    'can_complete_maintenance': 0,
    'can_view_warehouses': 1,
    'can_add_warehouse': 0,
    'can_edit_warehouse': 0,
    'can_delete_warehouse': 0,
    'can_view_reports': 1,
    'can_export_reports': 0,
    'can_manage_users': 0,
    'can_manage_backup': 0,
    'can_change_own_password': 1
}

# الألوان
COLORS = {
    'primary': '#3498db',
    'success': '#27ae60',
    'warning': '#f39c12',
    'danger': '#e74c3c',
    'info': '#00bcd4',
    'purple': '#9b59b6',
    'dark': '#2c3e50',
    'light': '#ecf0f1',
    'white': '#ffffff',
    'gray': '#95a5a6',
    'orange': '#e67e22',
    'teal': '#1abc9c'
}

# ================================ إدارة قاعدة البيانات ================================
class DatabaseManager:
    """مدير قاعدة البيانات - نمط Singleton"""
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.init_database()
        return cls._instance
    
    def init_database(self):
        """تهيئة قاعدة البيانات وإنشاء الجداول"""
        self.conn = sqlite3.connect(DB_NAME, check_same_thread=False)
        self.conn.execute("PRAGMA foreign_keys = ON")
        self.create_tables()
        self.init_default_data()
    
    @contextmanager
    def get_cursor(self):
        """إنشاء مؤشر قاعدة البيانات مع الإغلاق التلقائي"""
        cursor = self.conn.cursor()
        try:
            yield cursor
            self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            raise e
        finally:
            cursor.close()
    
    def create_tables(self):
        """إنشاء جداول قاعدة البيانات"""
        queries = [
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                full_name TEXT,
                role TEXT DEFAULT 'operator',
                is_active INTEGER DEFAULT 1,
                last_login DATETIME,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER,
                FOREIGN KEY (created_by) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS user_permissions (
                user_id INTEGER PRIMARY KEY,
                can_view_dashboard INTEGER DEFAULT 1,
                can_manage_carts INTEGER DEFAULT 1,
                can_add_cart INTEGER DEFAULT 1,
                can_edit_cart INTEGER DEFAULT 0,
                can_delete_cart INTEGER DEFAULT 0,
                can_move_cart INTEGER DEFAULT 1,
                can_view_movements INTEGER DEFAULT 1,
                can_manage_maintenance INTEGER DEFAULT 1,
                can_complete_maintenance INTEGER DEFAULT 0,
                can_view_warehouses INTEGER DEFAULT 1,
                can_add_warehouse INTEGER DEFAULT 0,
                can_edit_warehouse INTEGER DEFAULT 0,
                can_delete_warehouse INTEGER DEFAULT 0,
                can_view_reports INTEGER DEFAULT 1,
                can_export_reports INTEGER DEFAULT 0,
                can_manage_users INTEGER DEFAULT 0,
                can_manage_backup INTEGER DEFAULT 0,
                can_change_own_password INTEGER DEFAULT 1,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS app_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                setting_key TEXT UNIQUE NOT NULL,
                setting_value TEXT,
                description TEXT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_by INTEGER,
                FOREIGN KEY (updated_by) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS warehouses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                capacity INTEGER NOT NULL,
                current_count INTEGER DEFAULT 0,
                location_type TEXT,
                description TEXT,
                is_active INTEGER DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER,
                FOREIGN KEY (created_by) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS carts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                serial_number TEXT UNIQUE NOT NULL,
                status TEXT CHECK(status IN ('sound', 'needs_maintenance', 'damaged')) DEFAULT 'sound',
                current_warehouse_id INTEGER,
                last_updated DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER,
                notes TEXT,
                FOREIGN KEY (current_warehouse_id) REFERENCES warehouses (id) ON DELETE SET NULL,
                FOREIGN KEY (created_by) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cart_id INTEGER NOT NULL,
                from_warehouse_id INTEGER,
                to_warehouse_id INTEGER NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                user_id INTEGER,
                notes TEXT,
                FOREIGN KEY (cart_id) REFERENCES carts (id) ON DELETE CASCADE,
                FOREIGN KEY (from_warehouse_id) REFERENCES warehouses (id) ON DELETE SET NULL,
                FOREIGN KEY (to_warehouse_id) REFERENCES warehouses (id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS maintenance_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cart_id INTEGER NOT NULL,
                maintenance_type TEXT,
                status TEXT DEFAULT 'pending',
                description TEXT,
                entry_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                completion_date DATETIME,
                user_id INTEGER,
                completed_by INTEGER,
                cost REAL DEFAULT 0,
                FOREIGN KEY (cart_id) REFERENCES carts (id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE SET NULL,
                FOREIGN KEY (completed_by) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS backups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT,
                backup_type TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                user_id INTEGER,
                file_size INTEGER,
                file_path TEXT,
                mega_link TEXT,
                status TEXT DEFAULT 'completed',
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE SET NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS system_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT,
                description TEXT,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE SET NULL
            )
            """
        ]
        
        with self.get_cursor() as cursor:
            for query in queries:
                cursor.execute(query)
    
    def init_default_data(self):
        """إدخال البيانات الافتراضية"""
        with self.get_cursor() as cursor:
            # إضافة المستخدم الرئيسي
            cursor.execute("SELECT * FROM users WHERE username = ?", (DEFAULT_USER,))
            admin = cursor.fetchone()
            
            if not admin:
                cursor.execute(
                    """INSERT INTO users (username, password, full_name, role, is_active) 
                       VALUES (?, ?, ?, 'admin', 1)""",
                    (DEFAULT_USER, DEFAULT_PASSWORD, 'سعود آل سعود')
                )
                admin_id = cursor.lastrowid
                
                # إضافة صلاحيات المدير
                permissions = DEFAULT_PERMISSIONS.copy()
                for key in permissions:
                    permissions[key] = 1
                
                columns = ['user_id'] + list(permissions.keys())
                values = [admin_id] + list(permissions.values())
                placeholders = ','.join(['?' for _ in columns])
                
                cursor.execute(
                    f"INSERT INTO user_permissions ({','.join(columns)}) VALUES ({placeholders})",
                    values
                )
            
            # إضافة إعدادات التطبيق الافتراضية
            cursor.execute("SELECT * FROM app_settings WHERE setting_key = 'app_name'")
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, description) VALUES (?, ?, ?)",
                    ('app_name', APP_NAME, 'اسم التطبيق الرئيسي')
                )
            
            cursor.execute("SELECT * FROM app_settings WHERE setting_key = 'company_name'")
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, description) VALUES (?, ?, ?)",
                    ('company_name', 'الرئاسة العامة لشؤون المسجد الحرام والمسجد النبوي', 'اسم الجهة المشغلة')
                )
            
            # إضافة إعدادات MEGA من المتغيرات البيئية
            cursor.execute("SELECT * FROM app_settings WHERE setting_key = 'mega_email'")
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, description) VALUES (?, ?, ?)",
                    ('mega_email', MEGA_EMAIL, 'بريد MEGA للنسخ الاحتياطي السحابي')
                )
            
            cursor.execute("SELECT * FROM app_settings WHERE setting_key = 'mega_password'")
            if not cursor.fetchone():
                cursor.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, description) VALUES (?, ?, ?)",
                    ('mega_password', MEGA_PASSWORD, 'كلمة مرور MEGA للنسخ الاحتياطي السحابي')
                )
            
            # إضافة المستودعات الأساسية
            for wh in WAREHOUSES:
                cursor.execute("SELECT * FROM warehouses WHERE name = ?", (wh['name'],))
                if not cursor.fetchone():
                    cursor.execute(
                        """INSERT INTO warehouses 
                           (name, capacity, current_count, location_type, description, is_active) 
                           VALUES (?, ?, 0, ?, ?, 1)""",
                        (wh['name'], wh['capacity'], wh['type'], wh['description'])
                    )
    
    def get_app_setting(self, key, default=None):
        """الحصول على إعداد التطبيق"""
        result = self.execute_query(
            "SELECT setting_value FROM app_settings WHERE setting_key = ?",
            (key,)
        )
        return result[0][0] if result else default
    
    def update_app_setting(self, key, value, user_id=None):
        """تحديث إعداد التطبيق"""
        with self.get_cursor() as cursor:
            cursor.execute(
                """UPDATE app_settings 
                   SET setting_value = ?, updated_at = CURRENT_TIMESTAMP, updated_by = ? 
                   WHERE setting_key = ?""",
                (value, user_id, key)
            )
    
    def execute_query(self, query, params=()):
        """تنفيذ استعلام مع إرجاع النتائج"""
        with self.get_cursor() as cursor:
            cursor.execute(query, params)
            return cursor.fetchall()
    
    def execute_insert(self, query, params=()):
        """تنفيذ إدخال وإرجاع آخر ID"""
        with self.get_cursor() as cursor:
            cursor.execute(query, params)
            return cursor.lastrowid
    
    def get_warehouse_count(self, warehouse_id):
        """الحصول على عدد العربات في مستودع معين"""
        result = self.execute_query(
            "SELECT COUNT(*) FROM carts WHERE current_warehouse_id = ? AND status != 'damaged'",
            (warehouse_id,)
        )
        return result[0][0] if result else 0
    
    def update_warehouse_count(self, warehouse_id):
        """تحديث عدد العربات في المستودع"""
        count = self.get_warehouse_count(warehouse_id)
        with self.get_cursor() as cursor:
            cursor.execute(
                "UPDATE warehouses SET current_count = ? WHERE id = ?",
                (count, warehouse_id)
            )
    
    def get_all_warehouses(self):
        """الحصول على جميع المستودعات النشطة"""
        return self.execute_query(
            "SELECT id, name FROM warehouses WHERE is_active = 1 ORDER BY name"
        )
    
    def get_user_permissions(self, user_id):
        """الحصول على صلاحيات المستخدم"""
        result = self.execute_query(
            "SELECT * FROM user_permissions WHERE user_id = ?",
            (user_id,)
        )
        
        if result:
            columns = ['user_id', 'can_view_dashboard', 'can_manage_carts', 'can_add_cart',
                      'can_edit_cart', 'can_delete_cart', 'can_move_cart', 'can_view_movements',
                      'can_manage_maintenance', 'can_complete_maintenance', 'can_view_warehouses',
                      'can_add_warehouse', 'can_edit_warehouse', 'can_delete_warehouse',
                      'can_view_reports', 'can_export_reports', 'can_manage_users',
                      'can_manage_backup', 'can_change_own_password', 'updated_at']
            
            permissions = {}
            for i, col in enumerate(columns):
                permissions[col] = result[0][i]
            return permissions
        else:
            permissions = DEFAULT_PERMISSIONS.copy()
            permissions['user_id'] = user_id
            return permissions
    
    def update_user_permissions(self, user_id, permissions):
        """تحديث صلاحيات المستخدم"""
        with self.get_cursor() as cursor:
            cursor.execute("SELECT * FROM user_permissions WHERE user_id = ?", (user_id,))
            if cursor.fetchone():
                set_clause = ','.join([f"{key}=?" for key in permissions.keys()])
                values = list(permissions.values()) + [user_id]
                cursor.execute(
                    f"UPDATE user_permissions SET {set_clause}, updated_at=CURRENT_TIMESTAMP WHERE user_id=?",
                    values
                )
            else:
                columns = ['user_id'] + list(permissions.keys())
                values = [user_id] + list(permissions.values())
                placeholders = ','.join(['?' for _ in columns])
                cursor.execute(
                    f"INSERT INTO user_permissions ({','.join(columns)}) VALUES ({placeholders})",
                    values
                )
    
    def log_action(self, user_id, action, description):
        """تسجيل إجراء في سجل النظام"""
        try:
            self.execute_insert(
                "INSERT INTO system_logs (user_id, action, description) VALUES (?, ?, ?)",
                (user_id, action, description)
            )
        except:
            pass

# ================================ تطبيق Flet الرئيسي ================================
class CartsManagementApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.db = DatabaseManager()
        self.current_user = None
        self.current_permissions = None
        self.backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
        
        # إنشاء مجلد النسخ الاحتياطي إذا لم يكن موجوداً
        if not os.path.exists(self.backup_dir):
            os.makedirs(self.backup_dir)
        
        # إعدادات الصفحة
        app_name = self.db.get_app_setting('app_name', APP_NAME)
        self.page.title = app_name
        self.page.rtl = True
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 0
        self.page.window_width = 1300
        self.page.window_height = 800
        self.page.window_min_width = 1000
        self.page.window_min_height = 600
        self.page.scroll = ft.ScrollMode.AUTO
        
        # المتغيرات العامة
        self.content_column = None
        self.sidebar = None
        
        # متغيرات البحث والفلترة
        self.cart_search_field = None
        self.cart_table = None
        self.movement_search_field = None
        self.movement_table = None
        self.maintenance_search_field = None
        self.maintenance_table = None
        self.warehouse_search_field = None
        self.warehouse_table = None
        self.user_search_field = None
        self.user_table = None
        
        # متغيرات التقارير
        self.report_type_dropdown = None
        self.period_dropdown = None
        self.preview_table = None
        
        # متغيرات النسخ الاحتياطي
        self.backup_progress = None
        self.backup_status = None
        self.backup_tree = None
        
        # متغيرات إعدادات MEGA
        self.mega_status_label = None
        
        # عرض شاشة تسجيل الدخول
        self.show_login_screen()
    
    # ================================ دوال مساعدة ================================
    def show_snack_bar(self, message, color=COLORS['success']):
        """عرض رسالة منبثقة"""
        self.page.snack_bar = ft.SnackBar(
            content=ft.Text(message, color=COLORS['white'], text_align=ft.TextAlign.RIGHT),
            bgcolor=color,
            show_close_icon=True
        )
        self.page.snack_bar.open = True
        self.page.update()
    
    def check_permission(self, permission):
        """التحقق من صلاحية المستخدم"""
        if not self.current_permissions:
            return False
        if self.current_user and self.current_user['role'] == 'admin':
            return True
        return self.current_permissions.get(permission, 0) == 1
    
    def clear_content(self):
        """مسح منطقة المحتوى"""
        if self.content_column:
            self.content_column.controls.clear()
            self.page.update()
    
    def show_loading(self):
        """عرض مؤشر تحميل"""
        return ft.Container(
            content=ft.Column([
                ft.ProgressRing(),
                ft.Text("جاري التحميل...", size=16, color=COLORS['gray'])
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            alignment=ft.alignment.center,
            expand=True
        )
    
    # ================================ شاشة تسجيل الدخول ================================
    def show_login_screen(self):
        """عرض شاشة تسجيل الدخول"""
        self.page.clean()
        
        app_name = self.db.get_app_setting('app_name', APP_NAME)
        company_name = self.db.get_app_setting('company_name', 'الرئاسة العامة لشؤون المسجد الحرام والمسجد النبوي')
        
        # إنشاء حقول الإدخال مع المراجع
        username_field = ft.TextField(
            hint_text="أدخل اسم المستخدم",
            border_radius=8,
            text_align=ft.TextAlign.RIGHT,
            bgcolor=COLORS['white'],
            border_color=COLORS['gray'],
            focused_border_color=COLORS['primary'],
            width=300,
            height=45,
            text_size=14,
            ref=ft.Ref[ft.TextField]()
        )
        
        password_field = ft.TextField(
            hint_text="أدخل كلمة المرور",
            password=True,
            can_reveal_password=True,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT,
            bgcolor=COLORS['white'],
            border_color=COLORS['gray'],
            focused_border_color=COLORS['primary'],
            width=300,
            height=45,
            text_size=14,
            ref=ft.Ref[ft.TextField]()
        )
        
        # تخزين المراجع
        self.username_field = username_field
        self.password_field = password_field
        
        # بطاقة تسجيل الدخول
        login_card = ft.Container(
            width=500,
            height=500,
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=30,
            content=ft.Column(
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=15,
                controls=[
                    ft.Icon(name=ft.icons.LOCAL_SHIPPING, size=80, color=COLORS['dark']),
                    ft.Text(app_name, size=22, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                    ft.Text(company_name, size=14, color=COLORS['gray']),
                    ft.Divider(height=20, color=ft.colors.TRANSPARENT),
                    
                    ft.Container(
                        content=ft.Column([
                            ft.Text("اسم المستخدم:", size=14, weight=ft.FontWeight.W_500, 
                                   text_align=ft.TextAlign.RIGHT),
                            username_field,
                        ], spacing=5),
                    ),
                    
                    ft.Container(
                        content=ft.Column([
                            ft.Text("كلمة المرور:", size=14, weight=ft.FontWeight.W_500,
                                   text_align=ft.TextAlign.RIGHT),
                            password_field,
                        ], spacing=5),
                    ),
                    
                    ft.Container(height=10),
                    
                    ft.ElevatedButton(
                        text="تسجيل الدخول",
                        width=200,
                        height=45,
                        bgcolor=COLORS['success'],
                        color=COLORS['white'],
                        style=ft.ButtonStyle(
                            shape=ft.RoundedRectangleBorder(radius=8),
                        ),
                        on_click=self.handle_login
                    ),
                    
                    ft.Container(height=20),
                    ft.Text("جميع الحقوق محفوظة © 2025", size=12, color=COLORS['gray']),
                ]
            )
        )
        
        # الحاوية الرئيسية
        main_container = ft.Container(
            expand=True,
            bgcolor=COLORS['light'],
            alignment=ft.alignment.center,
            content=login_card
        )
        
        self.page.add(main_container)
        self.page.update()
    
    def handle_login(self, e):
        """معالجة تسجيل الدخول"""
        username = self.username_field.value.strip() if self.username_field.value else ""
        password = self.password_field.value.strip() if self.password_field.value else ""
        
        if not username or not password:
            self.show_snack_bar("الرجاء إدخال اسم المستخدم وكلمة المرور", COLORS['danger'])
            return
        
        result = self.db.execute_query(
            "SELECT id, username, role, is_active FROM users WHERE username = ? AND password = ?",
            (username, password)
        )
        
        if result:
            user_id, username, role, is_active = result[0]
            
            if not is_active:
                self.show_snack_bar("هذا المستخدم غير نشط. الرجاء التواصل مع المدير", COLORS['danger'])
                return
            
            self.current_user = {
                'id': user_id,
                'username': username,
                'role': role
            }
            
            self.db.execute_query(
                "UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = ?",
                (user_id,)
            )
            
            self.current_permissions = self.db.get_user_permissions(user_id)
            self.db.log_action(user_id, 'login', f'تسجيل دخول المستخدم {username}')
            self.show_main_screen()
        else:
            self.show_snack_bar("اسم المستخدم أو كلمة المرور غير صحيحة", COLORS['danger'])
    
    # ================================ الشاشة الرئيسية ================================
    def show_main_screen(self):
        """عرض الشاشة الرئيسية"""
        self.page.clean()
        
        # الصف الرئيسي
        main_row = ft.Row(
            spacing=0,
            controls=[]
        )
        
        # ===== الشريط الجانبي =====
        self.sidebar = ft.Container(
            width=280,
            height=self.page.window_height,
            bgcolor=COLORS['dark'],
            padding=ft.padding.only(top=20, bottom=20, right=20, left=20),
            content=ft.Column(
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=5,
                controls=[
                    # معلومات المستخدم
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(name=ft.icons.LOCAL_SHIPPING, size=60, color=COLORS['white']),
                            ft.Text(
                                self.db.get_app_setting('app_name', APP_NAME),
                                size=16,
                                weight=ft.FontWeight.BOLD,
                                color=COLORS['white'],
                                text_align=ft.TextAlign.CENTER
                            ),
                            ft.Text(f"مرحباً {self.current_user['username']}", 
                                   size=14, color=COLORS['gray']),
                            ft.Text(
                                "(مدير النظام)" if self.current_user['role'] == 'admin' else "",
                                size=12, 
                                color=COLORS['warning']
                            ),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER)
                    ),
                    
                    ft.Divider(color=COLORS['gray'], height=1),
                    
                    # قائمة التنقل
                    ft.Column(
                        spacing=2,
                        controls=self.build_menu_items(),
                        scroll=ft.ScrollMode.AUTO,
                    ),
                    
                    ft.Container(height=10),
                    ft.Divider(color=COLORS['gray'], height=1),
                    
                    # زر تسجيل الخروج
                    ft.Container(
                        margin=ft.margin.only(top=20),
                        content=ft.ElevatedButton(
                            text="تسجيل الخروج",
                            icon=ft.icons.LOGOUT,
                            width=240,
                            height=45,
                            bgcolor=COLORS['danger'],
                            color=COLORS['white'],
                            style=ft.ButtonStyle(
                                shape=ft.RoundedRectangleBorder(radius=8),
                            ),
                            on_click=self.logout
                        )
                    )
                ]
            )
        )
        
        # ===== منطقة المحتوى =====
        self.content_column = ft.Column(
            spacing=20,
            scroll=ft.ScrollMode.AUTO,
            expand=True
        )
        
        content_container = ft.Container(
            expand=True,
            bgcolor=COLORS['light'],
            padding=20,
            content=self.content_column
        )
        
        main_row.controls.extend([content_container, self.sidebar])
        self.page.add(main_row)
        self.page.update()
        
        # عرض لوحة التحكم بشكل افتراضي
        if self.check_permission('can_view_dashboard'):
            self.show_dashboard()
    
    def build_menu_items(self):
        """بناء عناصر القائمة"""
        menu_items = []
        
        menu_config = [
            ("📊", "لوحة التحكم", self.show_dashboard, 'can_view_dashboard'),
            ("🛒", "إدارة العربات", self.show_cart_management, 'can_manage_carts'),
            ("🔄", "حركة العربات", self.show_cart_movement, None, ['can_move_cart', 'can_view_movements']),
            ("🔧", "الصيانة", self.show_maintenance, 'can_manage_maintenance'),
            ("🏢", "المستودعات", self.show_warehouse_management, 'can_view_warehouses'),
            ("📈", "التقارير", self.show_reports, 'can_view_reports'),
        ]
        
        for icon, text, handler, perm, or_perms in [(*item, None) if len(item) == 4 else item for item in menu_config]:
            if perm:
                if self.check_permission(perm):
                    menu_items.append(self.create_menu_button(icon, text, handler))
            elif or_perms:
                if any(self.check_permission(p) for p in or_perms):
                    menu_items.append(self.create_menu_button(icon, text, handler))
        
        # عناصر المدير
        if self.current_user['role'] == 'admin':
            if self.check_permission('can_manage_users'):
                menu_items.append(self.create_menu_button("👥", "إدارة المستخدمين", self.show_user_management))
                menu_items.append(self.create_menu_button("⚙️", "إعدادات النظام", self.show_system_settings))
            
            if self.check_permission('can_manage_backup'):
                menu_items.append(self.create_menu_button("💾", "النسخ الاحتياطي", self.show_backup))
        
        # تغيير كلمة المرور
        if self.check_permission('can_change_own_password'):
            menu_items.append(self.create_menu_button("🔑", "تغيير كلمة المرور", self.show_change_password))
        
        return menu_items
    
    def create_menu_button(self, icon, text, on_click):
        """إنشاء زر قائمة"""
        return ft.Container(
            width=240,
            content=ft.TextButton(
                content=ft.Row([
                    ft.Text(f"{icon}  {text}", size=14, color=COLORS['white']),
                ], alignment=ft.MainAxisAlignment.START),
                style=ft.ButtonStyle(
                    color=COLORS['white'],
                    overlay_color=COLORS['primary'],
                    padding=ft.padding.symmetric(horizontal=15, vertical=10),
                ),
                on_click=lambda e: on_click()
            )
        )
    
    def logout(self, e):
        """تسجيل الخروج"""
        def confirm_logout(e):
            if self.current_user:
                self.db.log_action(self.current_user['id'], 'logout',
                                  f'تسجيل خروج المستخدم {self.current_user["username"]}')
            self.current_user = None
            self.current_permissions = None
            dialog.open = False
            self.page.update()
            self.show_login_screen()
        
        def cancel_logout(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تسجيل الخروج"),
            content=ft.Text("هل أنت متأكد من تسجيل الخروج؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_logout),
                ft.TextButton("لا", on_click=cancel_logout),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ================================ لوحة التحكم ================================
    def show_dashboard(self):
        """عرض لوحة التحكم"""
        if not self.check_permission('can_view_dashboard'):
            self.show_snack_bar("غير مصرح لك بعرض لوحة التحكم", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Container(
                content=ft.Row([
                    ft.Text("لوحة التحكم", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                    ft.Text(datetime.now().strftime('%Y-%m-%d %H:%M'), 
                           size=14, color=COLORS['gray']),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                margin=ft.margin.only(bottom=20)
            )
        )
        
        # جلب الإحصائيات
        total_carts = self.db.execute_query("SELECT COUNT(*) FROM carts")[0][0] or 0
        sound_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'sound'")[0][0] or 0
        maintenance_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'needs_maintenance'")[0][0] or 0
        damaged_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'damaged'")[0][0] or 0
        total_warehouses = self.db.execute_query("SELECT COUNT(*) FROM warehouses WHERE is_active = 1")[0][0] or 0
        total_movements = self.db.execute_query("SELECT COUNT(*) FROM movements")[0][0] or 0
        pending_maintenance = self.db.execute_query("SELECT COUNT(*) FROM maintenance_records WHERE status = 'pending'")[0][0] or 0
        total_users = self.db.execute_query("SELECT COUNT(*) FROM users WHERE is_active = 1")[0][0] or 0
        
        # بطاقات الإحصائيات - الصف الأول
        stats_row1 = ft.ResponsiveRow(
            spacing=10,
            controls=[
                self.create_stat_card("🚛", "إجمالي العربات", total_carts, COLORS['primary'],
                                     f"زيادة 12% عن الشهر الماضي", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("✅", "عربات سليمة", sound_carts, COLORS['success'],
                                     f"{sound_carts/total_carts*100:.1f}% من الإجمالي" if total_carts > 0 else "0%", 
                                     col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("🔧", "تحتاج صيانة", maintenance_carts, COLORS['warning'],
                                     f"{maintenance_carts/total_carts*100:.1f}% من الإجمالي" if total_carts > 0 else "0%", 
                                     col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("⚠️", "عربات تالفة", damaged_carts, COLORS['danger'],
                                     f"{damaged_carts/total_carts*100:.1f}% من الإجمالي" if total_carts > 0 else "0%", 
                                     col={"sm": 6, "md": 3, "lg": 3}),
            ]
        )
        
        # بطاقات الإحصائيات - الصف الثاني
        stats_row2 = ft.ResponsiveRow(
            spacing=10,
            controls=[
                self.create_stat_card("🏢", "المستودعات", total_warehouses, COLORS['purple'], 
                                     "مستودع نشط", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("🔄", "حركات اليوم", total_movements, COLORS['info'], 
                                     "آخر 24 ساعة", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("🔧", "بانتظار الصيانة", pending_maintenance, COLORS['orange'], 
                                     f"{pending_maintenance} عربة", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("👥", "المستخدمين", total_users, COLORS['teal'], 
                                     f"{total_users} مستخدم نشط", col={"sm": 6, "md": 3, "lg": 3}),
            ]
        )
        
        self.content_column.controls.append(stats_row1)
        self.content_column.controls.append(ft.Container(height=10))
        self.content_column.controls.append(stats_row2)
        self.content_column.controls.append(ft.Container(height=20))
        
        # حالة المستودعات وآخر الحركات
        charts_row = ft.ResponsiveRow(
            spacing=10,
            controls=[
                ft.Container(
                    col={"sm": 12, "md": 6, "lg": 6},
                    bgcolor=COLORS['white'],
                    border_radius=10,
                    border=ft.border.all(1, COLORS['gray']),
                    padding=15,
                    content=ft.Column([
                        ft.Text("حالة المستودعات", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                        ft.Divider(height=1, color=COLORS['light']),
                        ft.Column(
                            spacing=15,
                            controls=self.get_warehouse_status_cards()
                        )
                    ])
                ),
                
                ft.Container(
                    col={"sm": 12, "md": 6, "lg": 6},
                    bgcolor=COLORS['white'],
                    border_radius=10,
                    border=ft.border.all(1, COLORS['gray']),
                    padding=15,
                    content=ft.Column([
                        ft.Text("آخر الحركات", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                        ft.Divider(height=1, color=COLORS['light']),
                        ft.Column(
                            spacing=10,
                            controls=self.get_recent_movements()
                        )
                    ])
                )
            ]
        )
        
        self.content_column.controls.append(charts_row)
        self.page.update()
    
    def create_stat_card(self, icon, title, value, color, subtitle, col=None):
        """إنشاء بطاقة إحصائية"""
        card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=15,
            content=ft.Column([
                ft.Row([
                    ft.Text(icon, size=30),
                    ft.Text(title, size=14, color=COLORS['gray']),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Container(height=5),
                ft.Text(f"{value:,}", size=24, weight=ft.FontWeight.BOLD, color=color),
                ft.Text(subtitle, size=11, color=COLORS['gray']),
            ])
        )
        
        if col:
            card.col = col
        
        return card
    
    def get_warehouse_status_cards(self):
        """الحصول على بطاقات حالة المستودعات"""
        cards = []
        warehouses = self.db.execute_query(
            "SELECT name, capacity, current_count FROM warehouses WHERE is_active = 1 ORDER BY id LIMIT 5"
        )
        
        for wh in warehouses:
            name, capacity, current = wh
            percentage = (current / capacity * 100) if capacity > 0 else 0
            
            if percentage >= 90:
                color = COLORS['danger']
            elif percentage >= 70:
                color = COLORS['warning']
            else:
                color = COLORS['success']
            
            cards.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(name, size=14, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                            ft.Text(f"{percentage:.1f}%", size=14, weight=ft.FontWeight.BOLD, color=color),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Container(
                            height=8,
                            bgcolor=COLORS['light'],
                            border_radius=4,
                            content=ft.Container(
                                width=max(percentage * 2, 5),
                                height=8,
                                bgcolor=color,
                                border_radius=4,
                            )
                        ),
                        ft.Row([
                            ft.Text(f"{current} / {capacity}", size=12, color=COLORS['gray']),
                        ], alignment=ft.MainAxisAlignment.END),
                    ])
                )
            )
        
        return cards
    
    def get_recent_movements(self):
        """الحصول على آخر الحركات"""
        movements = []
        data = self.db.execute_query("""
            SELECT c.serial_number, w1.name, w2.name, m.timestamp
            FROM movements m
            JOIN carts c ON m.cart_id = c.id
            LEFT JOIN warehouses w1 ON m.from_warehouse_id = w1.id
            JOIN warehouses w2 ON m.to_warehouse_id = w2.id
            ORDER BY m.timestamp DESC
            LIMIT 8
        """)
        
        for m in data:
            serial, from_wh, to_wh, timestamp = m
            movements.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(f"🚛 {serial}", size=13, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                            ft.Text(timestamp[:16] if timestamp else "", size=11, color=COLORS['gray']),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Text(f"{from_wh or '—'}  ←  {to_wh}", size=12, color=COLORS['primary']),
                        ft.Divider(height=1, color=COLORS['light']),
                    ])
                )
            )
        
        return movements if movements else [ft.Text("لا توجد حركات", size=14, color=COLORS['gray'])]
    
    # ================================ إدارة العربات ================================
    def show_cart_management(self):
        """عرض صفحة إدارة العربات"""
        if not self.check_permission('can_manage_carts'):
            self.show_snack_bar("غير مصرح لك بإدارة العربات", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        title_row = ft.Row([
            ft.Text("إدارة العربات", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
            ft.Row([
                ft.TextField(
                    hint_text="بحث...",
                    width=250,
                    height=40,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    prefix=ft.Icon(ft.icons.SEARCH),
                    on_change=self.filter_carts,
                    ref=ft.Ref[ft.TextField]()
                ),
                ft.ElevatedButton(
                    text="إضافة عربة جديدة",
                    icon=ft.icons.ADD,
                    bgcolor=COLORS['success'],
                    color=COLORS['white'],
                    style=ft.ButtonStyle(
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=self.show_add_cart_dialog,
                    visible=self.check_permission('can_add_cart')
                ),
            ])
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
        
        self.content_column.controls.append(title_row)
        self.content_column.controls.append(ft.Container(height=20))
        
        # تخزين مرجع حقل البحث
        self.cart_search_field = title_row.controls[1].controls[0]
        
        # جدول العربات
        self.cart_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("الرقم", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الرقم التسلسلي", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الحالة", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("المستودع الحالي", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("آخر تحديث", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الإجراءات", size=14, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            horizontal_margin=10,
            column_spacing=30,
            heading_row_color=COLORS['light'],
            heading_row_height=50,
            data_row_max_height=50,
            expand=True
        )
        
        # حاوية الجدول مع التمرير
        table_container = ft.Container(
            content=ft.Column([
                self.cart_table
            ], scroll=ft.ScrollMode.AUTO),
            expand=True,
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=15
        )
        
        self.content_column.controls.append(table_container)
        self.load_carts()
        self.page.update()
    
    def load_carts(self):
        """تحميل قائمة العربات"""
        if not self.cart_table:
            return
        
        self.cart_table.rows.clear()
        
        carts = self.db.execute_query("""
            SELECT c.id, c.serial_number, c.status, w.name, c.last_updated
            FROM carts c
            LEFT JOIN warehouses w ON c.current_warehouse_id = w.id
            ORDER BY c.id DESC
        """)
        
        for cart in carts:
            cart_id, serial, status, warehouse, updated = cart
            status_text = CART_STATUS.get(status, status)
            
            # تحديد لون الصف حسب الحالة
            row_color = None
            if status == 'sound':
                row_color = ft.colors.with_opacity(0.1, COLORS['success'])
            elif status == 'needs_maintenance':
                row_color = ft.colors.with_opacity(0.1, COLORS['warning'])
            elif status == 'damaged':
                row_color = ft.colors.with_opacity(0.1, COLORS['danger'])
            
            # أزرار الإجراءات
            actions_row = ft.Row([
                ft.IconButton(
                    icon=ft.icons.EDIT,
                    icon_size=18,
                    icon_color=COLORS['primary'],
                    tooltip="تعديل",
                    on_click=lambda e, cid=cart_id, s=serial: self.edit_cart(cid, s),
                    visible=self.check_permission('can_edit_cart')
                ),
                ft.IconButton(
                    icon=ft.icons.DELETE,
                    icon_size=18,
                    icon_color=COLORS['danger'],
                    tooltip="حذف",
                    on_click=lambda e, cid=cart_id: self.delete_cart(cid),
                    visible=self.check_permission('can_delete_cart')
                ),
            ], spacing=5)
            
            self.cart_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(cart_id), size=13)),
                        ft.DataCell(ft.Text(serial, size=13)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(status_text, size=13, color=COLORS['white']),
                            bgcolor=COLORS['success'] if status == 'sound' else 
                                   COLORS['warning'] if status == 'needs_maintenance' else 
                                   COLORS['danger'],
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=4
                        )),
                        ft.DataCell(ft.Text(warehouse or "غير محدد", size=13)),
                        ft.DataCell(ft.Text(updated[:10] if updated else "", size=13)),
                        ft.DataCell(actions_row),
                    ],
                    color=row_color
                )
            )
        
        self.page.update()
    
    def filter_carts(self, e):
        """فلترة العربات حسب البحث"""
        if not self.cart_table:
            return
        
        search_text = e.control.value.strip().lower() if e.control.value else ""
        
        for row in self.cart_table.rows[:]:
            values = []
            for cell in row.cells[:4]:  # الأعمدة الأولى فقط
                if isinstance(cell.content, ft.Text):
                    values.append(cell.content.value.lower())
                elif isinstance(cell.content, ft.Container):
                    if isinstance(cell.content.content, ft.Text):
                        values.append(cell.content.content.value.lower())
            
            if search_text:
                if not any(search_text in val for val in values):
                    self.cart_table.rows.remove(row)
            else:
                # إعادة تحميل الجدول بالكامل إذا كان البحث فارغاً
                self.cart_table.rows.clear()
                self.load_carts()
                break
        
        self.page.update()
    
    def show_add_cart_dialog(self, e):
        """عرض نافذة إضافة عربة جديدة"""
        if not self.check_permission('can_add_cart'):
            self.show_snack_bar("غير مصرح لك بإضافة عربات جديدة", COLORS['danger'])
            return
        
        # جلب قائمة المستودعات
        warehouses = self.db.get_all_warehouses()
        warehouse_options = [w[1] for w in warehouses]
        
        # حقول الإدخال
        serial_field = ft.TextField(
            label="الرقم التسلسلي",
            width=300,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT,
            autofocus=True
        )
        
        status_dropdown = ft.Dropdown(
            label="الحالة",
            width=300,
            options=[
                ft.dropdown.Option("سليمة"),
                ft.dropdown.Option("تحتاج صيانة"),
                ft.dropdown.Option("تالفة"),
            ],
            value="سليمة"
        )
        
        warehouse_dropdown = ft.Dropdown(
            label="المستودع",
            width=300,
            options=[ft.dropdown.Option(name) for name in warehouse_options] if warehouse_options else [],
            value=warehouse_options[0] if warehouse_options else None
        )
        
        notes_field = ft.TextField(
            label="ملاحظات",
            width=300,
            multiline=True,
            min_lines=3,
            max_lines=5,
            text_align=ft.TextAlign.RIGHT
        )
        
        def save_cart(e):
            serial = serial_field.value.strip() if serial_field.value else ""
            status_text = status_dropdown.value
            warehouse_name = warehouse_dropdown.value
            notes = notes_field.value or ""
            
            if not serial:
                self.show_snack_bar("الرجاء إدخال الرقم التسلسلي", COLORS['danger'])
                return
            
            status_map = {
                "سليمة": 'sound',
                "تحتاج صيانة": 'needs_maintenance',
                "تالفة": 'damaged'
            }
            status = status_map.get(status_text, 'sound')
            
            warehouse_id = None
            for w in warehouses:
                if w[1] == warehouse_name:
                    warehouse_id = w[0]
                    break
            
            try:
                cart_id = self.db.execute_insert(
                    """INSERT INTO carts 
                       (serial_number, status, current_warehouse_id, created_by, notes) 
                       VALUES (?, ?, ?, ?, ?)""",
                    (serial, status, warehouse_id, self.current_user['id'], notes)
                )
                
                if warehouse_id:
                    self.db.update_warehouse_count(warehouse_id)
                
                self.db.log_action(self.current_user['id'], 'add_cart',
                                  f'إضافة عربة جديدة رقم {serial}')
                
                dialog.open = False
                self.page.update()
                self.show_snack_bar("تم إضافة العربة بنجاح", COLORS['success'])
                self.load_carts()
                
            except sqlite3.IntegrityError:
                self.show_snack_bar("الرقم التسلسلي موجود مسبقاً", COLORS['danger'])
        
        dialog = ft.AlertDialog(
            title=ft.Text("إضافة عربة جديدة", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    serial_field,
                    status_dropdown,
                    warehouse_dropdown,
                    notes_field,
                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_cart, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def edit_cart(self, cart_id, serial):
        """تعديل بيانات العربة"""
        if not self.check_permission('can_edit_cart'):
            self.show_snack_bar("غير مصرح لك بتعديل العربات", COLORS['danger'])
            return
        
        result = self.db.execute_query("""
            SELECT c.status, w.name, c.notes
            FROM carts c
            LEFT JOIN warehouses w ON c.current_warehouse_id = w.id
            WHERE c.id = ?
        """, (cart_id,))
        
        if not result:
            self.show_snack_bar("العربة غير موجودة", COLORS['danger'])
            return
        
        status, warehouse, notes = result[0]
        status_text = CART_STATUS.get(status, status)
        
        # جلب قائمة المستودعات
        warehouses = self.db.get_all_warehouses()
        warehouse_options = [w[1] for w in warehouses]
        
        # حقول الإدخال
        serial_display = ft.TextField(
            label="الرقم التسلسلي",
            width=300,
            value=serial,
            read_only=True,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT
        )
        
        status_dropdown = ft.Dropdown(
            label="الحالة",
            width=300,
            options=[
                ft.dropdown.Option("سليمة"),
                ft.dropdown.Option("تحتاج صيانة"),
                ft.dropdown.Option("تالفة"),
            ],
            value=status_text
        )
        
        warehouse_dropdown = ft.Dropdown(
            label="المستودع",
            width=300,
            options=[ft.dropdown.Option(name) for name in warehouse_options] if warehouse_options else [],
            value=warehouse or (warehouse_options[0] if warehouse_options else None)
        )
        
        notes_field = ft.TextField(
            label="ملاحظات",
            width=300,
            value=notes or "",
            multiline=True,
            min_lines=3,
            max_lines=5,
            text_align=ft.TextAlign.RIGHT
        )
        
        def save_edit(e):
            new_status_text = status_dropdown.value
            new_warehouse_name = warehouse_dropdown.value
            new_notes = notes_field.value or ""
            
            status_map = {
                "سليمة": 'sound',
                "تحتاج صيانة": 'needs_maintenance',
                "تالفة": 'damaged'
            }
            new_status = status_map.get(new_status_text, 'sound')
            
            new_warehouse_id = None
            for w in warehouses:
                if w[1] == new_warehouse_name:
                    new_warehouse_id = w[0]
                    break
            
            old_warehouse = self.db.execute_query(
                "SELECT current_warehouse_id FROM carts WHERE id = ?",
                (cart_id,)
            )[0][0]
            
            self.db.execute_query(
                """UPDATE carts 
                   SET status = ?, current_warehouse_id = ?, last_updated = CURRENT_TIMESTAMP, notes = ? 
                   WHERE id = ?""",
                (new_status, new_warehouse_id, new_notes, cart_id)
            )
            
            if old_warehouse:
                self.db.update_warehouse_count(old_warehouse)
            if new_warehouse_id:
                self.db.update_warehouse_count(new_warehouse_id)
            
            self.db.log_action(self.current_user['id'], 'edit_cart',
                              f'تعديل العربة رقم {serial}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تعديل العربة بنجاح", COLORS['success'])
            self.load_carts()
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تعديل العربة: {serial}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    serial_display,
                    status_dropdown,
                    warehouse_dropdown,
                    notes_field,
                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_edit, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def delete_cart(self, cart_id):
        """حذف عربة"""
        if not self.check_permission('can_delete_cart'):
            self.show_snack_bar("غير مصرح لك بحذف العربات", COLORS['danger'])
            return
        
        def confirm_delete(e):
            result = self.db.execute_query(
                "SELECT current_warehouse_id, serial_number FROM carts WHERE id = ?",
                (cart_id,)
            )
            
            if result:
                warehouse_id, serial = result[0]
                self.db.execute_query("DELETE FROM carts WHERE id = ?", (cart_id,))
                
                if warehouse_id:
                    self.db.update_warehouse_count(warehouse_id)
                
                self.db.log_action(self.current_user['id'], 'delete_cart',
                                  f'حذف العربة رقم {serial}')
                
                dialog.open = False
                self.page.update()
                self.show_snack_bar("تم حذف العربة بنجاح", COLORS['success'])
                self.load_carts()
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد الحذف"),
            content=ft.Text("هل أنت متأكد من حذف هذه العربة؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_delete),
                ft.TextButton("لا", on_click=cancel_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ================================ حركة العربات ================================
    def show_cart_movement(self):
        """عرض صفحة حركة العربات"""
        if not self.check_permission('can_move_cart') and not self.check_permission('can_view_movements'):
            self.show_snack_bar("غير مصرح لك بعرض حركة العربات", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("حركة العربات - نقل بين المستودعات", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== قسم نقل العربة =====
        if self.check_permission('can_move_cart'):
            # جلب البيانات
            carts = self.db.execute_query("""
                SELECT c.id, c.serial_number, w.name
                FROM carts c
                LEFT JOIN warehouses w ON c.current_warehouse_id = w.id
                WHERE c.current_warehouse_id IS NOT NULL AND c.status != 'damaged'
                ORDER BY c.serial_number
            """)
            
            warehouses = self.db.get_all_warehouses()
            warehouse_dict = {w[1]: w[0] for w in warehouses}
            warehouse_names = list(warehouse_dict.keys())
            
            cart_options = [f"{c[1]} - ({c[2]})" for c in carts]
            
            # حقول الإدخال
            cart_dropdown = ft.Dropdown(
                label="اختر العربة",
                width=350,
                options=[ft.dropdown.Option(opt) for opt in cart_options],
                on_change=lambda e: self.update_from_warehouse(e, carts)
            )
            
            from_warehouse_dropdown = ft.Dropdown(
                label="من مستودع",
                width=250,
                options=[ft.dropdown.Option(name) for name in warehouse_names],
            )
            
            to_warehouse_dropdown = ft.Dropdown(
                label="إلى مستودع",
                width=250,
                options=[ft.dropdown.Option(name) for name in warehouse_names],
            )
            
            notes_field = ft.TextField(
                label="ملاحظات",
                width=350,
                multiline=True,
                min_lines=2,
                max_lines=3,
                text_align=ft.TextAlign.RIGHT
            )
            
            # تخزين المراجع
            self.cart_dropdown = cart_dropdown
            self.from_warehouse_dropdown = from_warehouse_dropdown
            self.to_warehouse_dropdown = to_warehouse_dropdown
            self.movement_notes = notes_field
            self.carts_data = carts
            
            def move_cart(e):
                cart_text = cart_dropdown.value
                from_warehouse = from_warehouse_dropdown.value
                to_warehouse = to_warehouse_dropdown.value
                notes = notes_field.value or ""
                
                if not cart_text:
                    self.show_snack_bar("الرجاء اختيار عربة", COLORS['danger'])
                    return
                
                if not from_warehouse:
                    self.show_snack_bar("الرجاء تحديد المستودع المصدر", COLORS['danger'])
                    return
                
                if not to_warehouse:
                    self.show_snack_bar("الرجاء اختيار مستودع الوجهة", COLORS['danger'])
                    return
                
                if from_warehouse == to_warehouse:
                    self.show_snack_bar("المستودع المصدر والهدف متطابقان", COLORS['danger'])
                    return
                
                from_id = warehouse_dict.get(from_warehouse)
                to_id = warehouse_dict.get(to_warehouse)
                
                cart_id = None
                for c in carts:
                    if f"{c[1]} - ({c[2]})" == cart_text:
                        cart_id = c[0]
                        break
                
                if not cart_id:
                    self.show_snack_bar("العربة غير موجودة", COLORS['danger'])
                    return
                
                result = self.db.execute_query(
                    "SELECT current_warehouse_id FROM carts WHERE id = ?",
                    (cart_id,)
                )
                
                if not result or result[0][0] != from_id:
                    self.show_snack_bar("العربة ليست في المستودع المصدر المحدد", COLORS['danger'])
                    return
                
                self.db.execute_query(
                    "UPDATE carts SET current_warehouse_id = ?, last_updated = CURRENT_TIMESTAMP WHERE id = ?",
                    (to_id, cart_id)
                )
                
                self.db.execute_insert(
                    """INSERT INTO movements 
                       (cart_id, from_warehouse_id, to_warehouse_id, user_id, notes) 
                       VALUES (?, ?, ?, ?, ?)""",
                    (cart_id, from_id, to_id, self.current_user['id'], notes)
                )
                
                self.db.update_warehouse_count(from_id)
                self.db.update_warehouse_count(to_id)
                
                self.db.log_action(self.current_user['id'], 'move_cart',
                                  f'نقل العربة {cart_text} من {from_warehouse} إلى {to_warehouse}')
                
                self.show_snack_bar("تم نقل العربة بنجاح", COLORS['success'])
                self.show_cart_movement()  # إعادة تحميل الصفحة
            
            # بطاقة نقل العربة
            movement_card = ft.Container(
                bgcolor=COLORS['white'],
                border_radius=10,
                border=ft.border.all(1, COLORS['gray']),
                padding=20,
                content=ft.Column([
                    ft.Text("نقل عربة", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                    ft.Divider(height=1, color=COLORS['light']),
                    
                    ft.ResponsiveRow([
                        ft.Container(
                            col={"sm": 12, "md": 6, "lg": 4},
                            content=cart_dropdown
                        ),
                        ft.Container(
                            col={"sm": 12, "md": 6, "lg": 4},
                            content=ft.Row([from_warehouse_dropdown, to_warehouse_dropdown])
                        ),
                        ft.Container(
                            col={"sm": 12, "md": 12, "lg": 4},
                            content=notes_field
                        ),
                    ]),
                    
                    ft.Container(height=10),
                    
                    ft.ElevatedButton(
                        text="نقل العربة",
                        icon=ft.icons.SWAP_HORIZ,
                        bgcolor=COLORS['primary'],
                        color=COLORS['white'],
                        style=ft.ButtonStyle(
                            shape=ft.RoundedRectangleBorder(radius=8),
                            padding=ft.padding.symmetric(horizontal=30, vertical=15)
                        ),
                        on_click=move_cart
                    )
                ])
            )
            
            self.content_column.controls.append(movement_card)
            self.content_column.controls.append(ft.Container(height=20))
        
        # ===== سجل الحركات =====
        if self.check_permission('can_view_movements'):
            history_card = ft.Container(
                bgcolor=COLORS['white'],
                border_radius=10,
                border=ft.border.all(1, COLORS['gray']),
                padding=20,
                expand=True,
                content=ft.Column([
                    ft.Row([
                        ft.Text("سجل الحركات", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                        ft.TextField(
                            hint_text="بحث في الحركات...",
                            width=250,
                            height=40,
                            border_radius=8,
                            text_align=ft.TextAlign.RIGHT,
                            prefix=ft.Icon(ft.icons.SEARCH),
                            on_change=self.filter_movements,
                            ref=ft.Ref[ft.TextField]()
                        ),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    
                    ft.Divider(height=1, color=COLORS['light']),
                    
                    # جدول الحركات
                    ft.DataTable(
                        columns=[
                            ft.DataColumn(ft.Text("التاريخ", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("العربة", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("من", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("إلى", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("المستخدم", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("ملاحظات", size=13, weight=ft.FontWeight.BOLD)),
                            ft.DataColumn(ft.Text("الإجراءات", size=13, weight=ft.FontWeight.BOLD)),
                        ],
                        rows=[],
                        horizontal_margin=10,
                        column_spacing=20,
                        heading_row_color=COLORS['light'],
                        heading_row_height=40,
                        data_row_max_height=40,
                        expand=True,
                        ref=ft.Ref[ft.DataTable]()
                    )
                ], expand=True)
            )
            
            self.movement_table = history_card.content.controls[2]
            self.movement_search_field = history_card.content.controls[0].controls[1]
            
            self.content_column.controls.append(history_card)
            self.load_movements()
        
        self.page.update()
    
    def update_from_warehouse(self, e, carts):
        """تحديث حقل المستودع المصدر بناءً على اختيار العربة"""
        cart_text = e.control.value
        if cart_text:
            for c in carts:
                if f"{c[1]} - ({c[2]})" == cart_text:
                    warehouse_name = c[2]
                    if warehouse_name:
                        self.from_warehouse_dropdown.value = warehouse_name
                        self.page.update()
                    break
    
    def load_movements(self):
        """تحميل سجل الحركات"""
        if not self.movement_table:
            return
        
        self.movement_table.rows.clear()
        
        movements = self.db.execute_query("""
            SELECT 
                m.id,
                m.timestamp,
                c.serial_number,
                w1.name as from_name,
                w2.name as to_name,
                u.username,
                m.notes
            FROM movements m
            JOIN carts c ON m.cart_id = c.id
            LEFT JOIN warehouses w1 ON m.from_warehouse_id = w1.id
            JOIN warehouses w2 ON m.to_warehouse_id = w2.id
            LEFT JOIN users u ON m.user_id = u.id
            ORDER BY m.timestamp DESC
            LIMIT 200
        """)
        
        for m in movements:
            movement_id, timestamp, serial, from_wh, to_wh, username, notes = m
            
            actions_row = ft.Row([
                ft.IconButton(
                    icon=ft.icons.DELETE,
                    icon_size=18,
                    icon_color=COLORS['danger'],
                    tooltip="حذف",
                    on_click=lambda e, mid=movement_id: self.delete_movement(mid),
                    visible=self.check_permission('can_delete_cart')
                ),
            ], spacing=5)
            
            self.movement_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(timestamp[:16] if timestamp else "", size=12)),
                        ft.DataCell(ft.Text(serial, size=12)),
                        ft.DataCell(ft.Text(from_wh or "-", size=12)),
                        ft.DataCell(ft.Text(to_wh, size=12)),
                        ft.DataCell(ft.Text(username or "", size=12)),
                        ft.DataCell(ft.Text((notes[:20] + '...') if notes and len(notes) > 20 else (notes or ""), size=12)),
                        ft.DataCell(actions_row),
                    ]
                )
            )
        
        self.page.update()
    
    def filter_movements(self, e):
        """فلترة سجل الحركات"""
        if not self.movement_table:
            return
        
        search_text = e.control.value.strip().lower() if e.control.value else ""
        
        # إعادة تحميل البيانات
        self.load_movements()
        
        if search_text:
            for row in self.movement_table.rows[:]:
                match = False
                for i, cell in enumerate(row.cells[:5]):  # الأعمدة الأولى
                    if isinstance(cell.content, ft.Text):
                        if search_text in cell.content.value.lower():
                            match = True
                            break
                
                if not match:
                    self.movement_table.rows.remove(row)
        
        self.page.update()
    
    def delete_movement(self, movement_id):
        """حذف حركة"""
        def confirm_delete(e):
            self.db.execute_query("DELETE FROM movements WHERE id = ?", (movement_id,))
            self.db.log_action(self.current_user['id'], 'delete_movement',
                              f'حذف حركة رقم {movement_id}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم حذف الحركة بنجاح", COLORS['success'])
            self.load_movements()
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد الحذف"),
            content=ft.Text("هل أنت متأكد من حذف هذه الحركة؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_delete),
                ft.TextButton("لا", on_click=cancel_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ================================ إدارة الصيانة ================================
    def show_maintenance(self):
        """عرض صفحة الصيانة"""
        if not self.check_permission('can_manage_maintenance'):
            self.show_snack_bar("غير مصرح لك بإدارة الصيانة", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("إدارة الصيانة", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # جلب البيانات
        carts = self.db.execute_query("""
            SELECT c.id, c.serial_number, w.name 
            FROM carts c
            LEFT JOIN warehouses w ON c.current_warehouse_id = w.id
            WHERE c.status != 'damaged'
            ORDER BY c.serial_number
        """)
        
        cart_options = [f"{c[1]} - ({c[2] or 'غير محدد'})" for c in carts]
        
        # ===== إحصائيات الصيانة =====
        pending = self.db.execute_query(
            "SELECT COUNT(*) FROM maintenance_records WHERE status = 'pending'"
        )[0][0] or 0
        
        in_progress = self.db.execute_query(
            "SELECT COUNT(*) FROM maintenance_records WHERE status = 'in_progress'"
        )[0][0] or 0
        
        completed = self.db.execute_query(
            "SELECT COUNT(*) FROM maintenance_records WHERE status = 'completed'"
        )[0][0] or 0
        
        total_cost = self.db.execute_query(
            "SELECT SUM(cost) FROM maintenance_records WHERE status = 'completed'"
        )[0][0] or 0
        
        # بطاقات الإحصائيات
        stats_row = ft.ResponsiveRow(
            spacing=10,
            controls=[
                self.create_stat_card("📋", "بانتظار الصيانة", pending, COLORS['warning'], 
                                     f"{pending} عربة", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("🔧", "قيد التنفيذ", in_progress, COLORS['primary'], 
                                     f"{in_progress} عربة", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("✅", "منجزة", completed, COLORS['success'], 
                                     f"{completed} عربة", col={"sm": 6, "md": 3, "lg": 3}),
                self.create_stat_card("💰", "إجمالي التكاليف", f"{total_cost:.0f} ر.س", COLORS['purple'], 
                                     "تكاليف الصيانة", col={"sm": 6, "md": 3, "lg": 3}),
            ]
        )
        
        self.content_column.controls.append(stats_row)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== إدخال عربية للصيانة =====
        input_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("إدخال عربية للصيانة", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.ResponsiveRow([
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 3},
                        content=ft.Dropdown(
                            label="العربة",
                            options=[ft.dropdown.Option(opt) for opt in cart_options],
                            ref=ft.Ref[ft.Dropdown]()
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 2},
                        content=ft.Dropdown(
                            label="نوع الصيانة",
                            options=[
                                ft.dropdown.Option("صيانة دورية"),
                                ft.dropdown.Option("إصلاح عطل"),
                                ft.dropdown.Option("تأهيل كامل"),
                                ft.dropdown.Option("فحص"),
                            ],
                            value="صيانة دورية"
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 2},
                        content=ft.Dropdown(
                            label="الحالة",
                            options=[
                                ft.dropdown.Option("تحتاج صيانة"),
                                ft.dropdown.Option("تالفة"),
                            ],
                            value="تحتاج صيانة"
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 2},
                        content=ft.TextField(
                            label="التكلفة",
                            value="0",
                            keyboard_type=ft.KeyboardType.NUMBER,
                            text_align=ft.TextAlign.RIGHT
                        )
                    ),
                ]),
                
                ft.ResponsiveRow([
                    ft.Container(
                        col={"sm": 12, "md": 12, "lg": 9},
                        content=ft.TextField(
                            label="وصف المشكلة",
                            multiline=True,
                            min_lines=2,
                            max_lines=3,
                            text_align=ft.TextAlign.RIGHT
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 12, "lg": 3},
                        content=ft.ElevatedButton(
                            text="إدخال للصيانة",
                            icon=ft.icons.BUILD,
                            bgcolor=COLORS['warning'],
                            color=COLORS['white'],
                            style=ft.ButtonStyle(
                                shape=ft.RoundedRectangleBorder(radius=8),
                                padding=ft.padding.symmetric(horizontal=20, vertical=15)
                            ),
                            on_click=lambda e: self.submit_maintenance(
                                e, carts, self.maintenance_inputs
                            )
                        )
                    ),
                ])
            ])
        )
        
        # تخزين مراجع حقول الإدخال
        self.maintenance_inputs = {
            'cart': input_card.content.controls[2].controls[0].content,
            'type': input_card.content.controls[2].controls[1].content,
            'status': input_card.content.controls[2].controls[2].content,
            'cost': input_card.content.controls[2].controls[3].content,
            'description': input_card.content.controls[3].controls[0].content
        }
        
        self.content_column.controls.append(input_card)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== سجل الصيانة =====
        records_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            expand=True,
            content=ft.Column([
                ft.Row([
                    ft.Text("سجل الصيانة", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                    ft.TextField(
                        hint_text="بحث في الصيانة...",
                        width=250,
                        height=40,
                        border_radius=8,
                        text_align=ft.TextAlign.RIGHT,
                        prefix=ft.Icon(ft.icons.SEARCH),
                        on_change=self.filter_maintenance,
                        ref=ft.Ref[ft.TextField]()
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                
                ft.Divider(height=1, color=COLORS['light']),
                
                # جدول الصيانة
                ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("التاريخ", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("العربة", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("نوع الصيانة", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("الحالة", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("الوصف", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("التكلفة", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("تاريخ الإنجاز", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("الإجراءات", size=13, weight=ft.FontWeight.BOLD)),
                    ],
                    rows=[],
                    horizontal_margin=10,
                    column_spacing=15,
                    heading_row_color=COLORS['light'],
                    heading_row_height=40,
                    data_row_max_height=50,
                    expand=True,
                    ref=ft.Ref[ft.DataTable]()
                )
            ], expand=True)
        )
        
        self.maintenance_table = records_card.content.controls[2]
        self.maintenance_search_field = records_card.content.controls[0].controls[1]
        
        self.content_column.controls.append(records_card)
        self.load_maintenance_records()
        self.page.update()
    
    def submit_maintenance(self, e, carts, inputs):
        """إدخال عربية للصيانة"""
        cart_text = inputs['cart'].value
        maint_type = inputs['type'].value
        status_text = inputs['status'].value
        cost_text = inputs['cost'].value
        description = inputs['description'].value or ""
        
        if not cart_text:
            self.show_snack_bar("الرجاء اختيار عربة", COLORS['danger'])
            return
        
        try:
            cost = float(cost_text or 0)
        except ValueError:
            cost = 0
        
        cart_id = None
        for c in carts:
            if f"{c[1]} - ({c[2] or 'غير محدد'})" == cart_text:
                cart_id = c[0]
                break
        
        if not cart_id:
            self.show_snack_bar("العربة غير موجودة", COLORS['danger'])
            return
        
        status_map = {
            "تحتاج صيانة": "needs_maintenance",
            "تالفة": "damaged"
        }
        new_status = status_map.get(status_text, "needs_maintenance")
        
        try:
            self.db.execute_query(
                "UPDATE carts SET status = ?, last_updated = CURRENT_TIMESTAMP WHERE id = ?",
                (new_status, cart_id)
            )
            
            self.db.execute_insert(
                """INSERT INTO maintenance_records 
                   (cart_id, maintenance_type, status, description, user_id, cost) 
                   VALUES (?, ?, 'pending', ?, ?, ?)""",
                (cart_id, maint_type, description, self.current_user['id'], cost)
            )
            
            self.db.log_action(self.current_user['id'], 'add_maintenance',
                              f'إدخال العربة {cart_text} للصيانة')
            
            self.show_snack_bar("تم إدخال العربة للصيانة", COLORS['success'])
            self.show_maintenance()  # إعادة تحميل الصفحة
            
        except Exception as e:
            self.show_snack_bar(f"حدث خطأ: {str(e)}", COLORS['danger'])
    
    def load_maintenance_records(self):
        """تحميل سجل الصيانة"""
        if not self.maintenance_table:
            return
        
        self.maintenance_table.rows.clear()
        
        records = self.db.execute_query("""
            SELECT 
                m.id,
                m.entry_date,
                c.serial_number,
                m.maintenance_type,
                m.status,
                m.description,
                m.cost,
                m.completion_date
            FROM maintenance_records m
            JOIN carts c ON m.cart_id = c.id
            ORDER BY m.entry_date DESC
            LIMIT 200
        """)
        
        for record in records:
            rec_id, entry_date, serial, maint_type, status, desc, cost, comp_date = record
            status_text = MAINTENANCE_STATUS.get(status, status)
            
            # تحديد لون الحالة
            status_color = COLORS['warning'] if status == 'pending' else \
                          COLORS['primary'] if status == 'in_progress' else \
                          COLORS['success']
            
            # أزرار الإجراءات
            actions_row = ft.Row(spacing=5)
            
            if status == 'pending' and self.check_permission('can_complete_maintenance'):
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.CHECK_CIRCLE,
                        icon_size=18,
                        icon_color=COLORS['success'],
                        tooltip="إتمام الصيانة",
                        on_click=lambda e, rid=rec_id: self.complete_maintenance(rid)
                    )
                )
            
            if self.check_permission('can_edit_cart'):
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.EDIT,
                        icon_size=18,
                        icon_color=COLORS['primary'],
                        tooltip="تعديل",
                        on_click=lambda e, rid=rec_id: self.edit_maintenance_record(rid)
                    )
                )
            
            if self.check_permission('can_delete_cart'):
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.DELETE,
                        icon_size=18,
                        icon_color=COLORS['danger'],
                        tooltip="حذف",
                        on_click=lambda e, rid=rec_id: self.delete_maintenance_record(rid)
                    )
                )
            
            self.maintenance_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(entry_date[:16] if entry_date else "", size=12)),
                        ft.DataCell(ft.Text(serial, size=12)),
                        ft.DataCell(ft.Text(maint_type, size=12)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(status_text, size=12, color=COLORS['white']),
                            bgcolor=status_color,
                            padding=ft.padding.symmetric(horizontal=8, vertical=2),
                            border_radius=4
                        )),
                        ft.DataCell(ft.Text((desc[:30] + '...') if desc and len(desc) > 30 else (desc or ""), size=12)),
                        ft.DataCell(ft.Text(f"{cost:.0f} ر.س", size=12)),
                        ft.DataCell(ft.Text(comp_date[:10] if comp_date else "", size=12)),
                        ft.DataCell(actions_row),
                    ]
                )
            )
        
        self.page.update()
    
    def complete_maintenance(self, record_id):
        """إتمام الصيانة"""
        if not self.check_permission('can_complete_maintenance'):
            self.show_snack_bar("غير مصرح لك بإتمام الصيانة", COLORS['danger'])
            return
        
        def confirm_complete(e):
            self.db.execute_query(
                """UPDATE maintenance_records 
                   SET status = 'completed', completion_date = CURRENT_TIMESTAMP, completed_by = ? 
                   WHERE id = ?""",
                (self.current_user['id'], record_id)
            )
            
            result = self.db.execute_query(
                "SELECT cart_id FROM maintenance_records WHERE id = ?",
                (record_id,)
            )
            
            if result:
                cart_id = result[0][0]
                self.db.execute_query(
                    "UPDATE carts SET status = 'sound', last_updated = CURRENT_TIMESTAMP WHERE id = ?",
                    (cart_id,)
                )
            
            self.db.log_action(self.current_user['id'], 'complete_maintenance',
                              f'إتمام صيانة للسجل رقم {record_id}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم إتمام الصيانة", COLORS['success'])
            self.load_maintenance_records()
        
        def cancel_complete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد إتمام الصيانة"),
            content=ft.Text("هل أنت متأكد من إتمام هذه الصيانة؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_complete),
                ft.TextButton("لا", on_click=cancel_complete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def edit_maintenance_record(self, record_id):
        """تعديل سجل الصيانة"""
        if not self.check_permission('can_edit_cart'):
            self.show_snack_bar("غير مصرح لك بتعديل سجلات الصيانة", COLORS['danger'])
            return
        
        result = self.db.execute_query("""
            SELECT m.cart_id, c.serial_number, m.maintenance_type, m.description, m.cost, m.status
            FROM maintenance_records m
            JOIN carts c ON m.cart_id = c.id
            WHERE m.id = ?
        """, (record_id,))
        
        if not result:
            self.show_snack_bar("سجل الصيانة غير موجود", COLORS['danger'])
            return
        
        cart_id, serial, maint_type, description, cost, status = result[0]
        
        # حقول الإدخال
        type_dropdown = ft.Dropdown(
            label="نوع الصيانة",
            width=300,
            options=[
                ft.dropdown.Option("صيانة دورية"),
                ft.dropdown.Option("إصلاح عطل"),
                ft.dropdown.Option("تأهيل كامل"),
                ft.dropdown.Option("فحص"),
            ],
            value=maint_type
        )
        
        status_dropdown = ft.Dropdown(
            label="الحالة",
            width=300,
            options=[
                ft.dropdown.Option("بانتظار الصيانة"),
                ft.dropdown.Option("قيد التنفيذ"),
                ft.dropdown.Option("منجزة"),
            ],
            value=MAINTENANCE_STATUS.get(status, status)
        )
        
        desc_field = ft.TextField(
            label="وصف المشكلة",
            width=300,
            value=description or "",
            multiline=True,
            min_lines=3,
            max_lines=5,
            text_align=ft.TextAlign.RIGHT
        )
        
        cost_field = ft.TextField(
            label="التكلفة",
            width=300,
            value=str(cost or 0),
            keyboard_type=ft.KeyboardType.NUMBER,
            text_align=ft.TextAlign.RIGHT
        )
        
        def save_edit(e):
            new_maint_type = type_dropdown.value
            new_status_text = status_dropdown.value
            new_description = desc_field.value or ""
            
            try:
                new_cost = float(cost_field.value or 0)
            except ValueError:
                new_cost = 0
            
            status_map = {
                "بانتظار الصيانة": "pending",
                "قيد التنفيذ": "in_progress",
                "منجزة": "completed"
            }
            new_status = status_map.get(new_status_text, "pending")
            
            self.db.execute_query(
                """UPDATE maintenance_records 
                   SET maintenance_type = ?, status = ?, description = ?, cost = ? 
                   WHERE id = ?""",
                (new_maint_type, new_status, new_description, new_cost, record_id)
            )
            
            if new_status == 'completed' and status != 'completed':
                self.db.execute_query(
                    "UPDATE carts SET status = 'sound', last_updated = CURRENT_TIMESTAMP WHERE id = ?",
                    (cart_id,)
                )
                self.db.execute_query(
                    "UPDATE maintenance_records SET completion_date = CURRENT_TIMESTAMP WHERE id = ?",
                    (record_id,)
                )
            
            self.db.log_action(self.current_user['id'], 'edit_maintenance',
                              f'تعديل سجل صيانة رقم {record_id}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تحديث سجل الصيانة بنجاح", COLORS['success'])
            self.load_maintenance_records()
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تعديل سجل الصيانة - {serial}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    type_dropdown,
                    status_dropdown,
                    desc_field,
                    cost_field,
                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_edit, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def delete_maintenance_record(self, record_id):
        """حذف سجل صيانة"""
        def confirm_delete(e):
            self.db.execute_query("DELETE FROM maintenance_records WHERE id = ?", (record_id,))
            self.db.log_action(self.current_user['id'], 'delete_maintenance',
                              f'حذف سجل صيانة رقم {record_id}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم حذف سجل الصيانة بنجاح", COLORS['success'])
            self.load_maintenance_records()
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد الحذف"),
            content=ft.Text("هل أنت متأكد من حذف سجل الصيانة هذا؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_delete),
                ft.TextButton("لا", on_click=cancel_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def filter_maintenance(self, e):
        """فلترة سجل الصيانة"""
        if not self.maintenance_table:
            return
        
        search_text = e.control.value.strip().lower() if e.control.value else ""
        
        # إعادة تحميل البيانات
        self.load_maintenance_records()
        
        if search_text:
            for row in self.maintenance_table.rows[:]:
                match = False
                for i, cell in enumerate(row.cells[:5]):  # الأعمدة الأولى
                    if isinstance(cell.content, ft.Text):
                        if search_text in cell.content.value.lower():
                            match = True
                            break
                    elif isinstance(cell.content, ft.Container):
                        if isinstance(cell.content.content, ft.Text):
                            if search_text in cell.content.content.value.lower():
                                match = True
                                break
                
                if not match:
                    self.maintenance_table.rows.remove(row)
        
        self.page.update()
    
    # ================================ إدارة المستودعات ================================
    def show_warehouse_management(self):
        """عرض صفحة إدارة المستودعات"""
        if not self.check_permission('can_view_warehouses'):
            self.show_snack_bar("غير مصرح لك بعرض المستودعات", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        title_row = ft.Row([
            ft.Text("إدارة المستودعات", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
            ft.Row([
                ft.TextField(
                    hint_text="بحث...",
                    width=250,
                    height=40,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    prefix=ft.Icon(ft.icons.SEARCH),
                    on_change=self.filter_warehouses,
                    ref=ft.Ref[ft.TextField]()
                ),
                ft.ElevatedButton(
                    text="إضافة مستودع",
                    icon=ft.icons.ADD_BUSINESS,
                    bgcolor=COLORS['success'],
                    color=COLORS['white'],
                    style=ft.ButtonStyle(
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=self.show_add_warehouse_dialog,
                    visible=self.check_permission('can_add_warehouse')
                ),
            ])
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
        
        self.content_column.controls.append(title_row)
        self.content_column.controls.append(ft.Container(height=20))
        
        # تخزين مرجع حقل البحث
        self.warehouse_search_field = title_row.controls[1].controls[0]
        
        # جدول المستودعات
        self.warehouse_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("المعرف", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("اسم المستودع", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("السعة", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("العدد الحالي", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("نسبة الإشغال", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الإجراءات", size=14, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            horizontal_margin=10,
            column_spacing=30,
            heading_row_color=COLORS['light'],
            heading_row_height=50,
            data_row_max_height=50,
            expand=True
        )
        
        # حاوية الجدول مع التمرير
        table_container = ft.Container(
            content=ft.Column([
                self.warehouse_table
            ], scroll=ft.ScrollMode.AUTO),
            expand=True,
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=15
        )
        
        self.content_column.controls.append(table_container)
        self.load_warehouses()
        self.page.update()
    
    def load_warehouses(self):
        """تحميل قائمة المستودعات"""
        if not self.warehouse_table:
            return
        
        self.warehouse_table.rows.clear()
        
        warehouses = self.db.execute_query("""
            SELECT id, name, capacity, current_count 
            FROM warehouses 
            WHERE is_active = 1
            ORDER BY id
        """)
        
        base_warehouse_names = [wh['name'] for wh in WAREHOUSES]
        
        for w in warehouses:
            wid, name, capacity, current = w
            percentage = (current / capacity * 100) if capacity > 0 else 0
            
            # تحديد لون نسبة الإشغال
            if percentage >= 90:
                color = COLORS['danger']
            elif percentage >= 70:
                color = COLORS['warning']
            else:
                color = COLORS['success']
            
            # أزرار الإجراءات
            actions_row = ft.Row(spacing=5)
            
            if self.check_permission('can_edit_warehouse'):
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.EDIT,
                        icon_size=18,
                        icon_color=COLORS['primary'],
                        tooltip="تعديل",
                        on_click=lambda e, wid=wid, n=name: self.edit_warehouse(wid, n)
                    )
                )
            
            if self.check_permission('can_delete_warehouse') and name not in base_warehouse_names:
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.DELETE,
                        icon_size=18,
                        icon_color=COLORS['danger'],
                        tooltip="حذف",
                        on_click=lambda e, wid=wid, n=name: self.delete_warehouse(wid, n)
                    )
                )
            
            self.warehouse_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(wid), size=13)),
                        ft.DataCell(ft.Text(name, size=13)),
                        ft.DataCell(ft.Text(str(capacity), size=13)),
                        ft.DataCell(ft.Text(str(current), size=13)),
                        ft.DataCell(ft.Container(
                            content=ft.Row([
                                ft.ProgressBar(
                                    width=80,
                                    value=percentage/100,
                                    bgcolor=COLORS['light'],
                                    color=color,
                                ),
                                ft.Text(f"{percentage:.1f}%", size=12, color=color),
                            ]),
                        )),
                        ft.DataCell(actions_row),
                    ]
                )
            )
        
        self.page.update()
    
    def filter_warehouses(self, e):
        """فلترة المستودعات حسب البحث"""
        if not self.warehouse_table:
            return
        
        search_text = e.control.value.strip().lower() if e.control.value else ""
        
        for row in self.warehouse_table.rows[:]:
            name_cell = row.cells[1].content
            if isinstance(name_cell, ft.Text):
                if search_text and search_text not in name_cell.value.lower():
                    self.warehouse_table.rows.remove(row)
        
        self.page.update()
    
    def show_add_warehouse_dialog(self, e):
        """عرض نافذة إضافة مستودع جديد"""
        if not self.check_permission('can_add_warehouse'):
            self.show_snack_bar("غير مصرح لك بإضافة مستودعات", COLORS['danger'])
            return
        
        # حقول الإدخال
        name_field = ft.TextField(
            label="اسم المستودع",
            width=300,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT,
            autofocus=True
        )
        
        capacity_field = ft.TextField(
            label="السعة",
            width=300,
            value="100",
            keyboard_type=ft.KeyboardType.NUMBER,
            text_align=ft.TextAlign.RIGHT
        )
        
        desc_field = ft.TextField(
            label="الوصف",
            width=300,
            text_align=ft.TextAlign.RIGHT
        )
        
        type_dropdown = ft.Dropdown(
            label="نوع المستودع",
            width=300,
            options=[
                ft.dropdown.Option("main", "رئيسي"),
                ft.dropdown.Option("external", "خارجي"),
                ft.dropdown.Option("north", "شمالي"),
                ft.dropdown.Option("south", "جنوبي"),
                ft.dropdown.Option("other", "آخر"),
            ],
            value="other"
        )
        
        def save_warehouse(e):
            name = name_field.value.strip() if name_field.value else ""
            capacity_text = capacity_field.value.strip() if capacity_field.value else ""
            description = desc_field.value or ""
            location_type = type_dropdown.value
            
            if not name:
                self.show_snack_bar("الرجاء إدخال اسم المستودع", COLORS['danger'])
                return
            
            try:
                capacity = int(capacity_text) if capacity_text else 100
            except ValueError:
                capacity = 100
            
            try:
                self.db.execute_insert(
                    """INSERT INTO warehouses 
                       (name, capacity, current_count, description, location_type, is_active, created_by) 
                       VALUES (?, ?, 0, ?, ?, 1, ?)""",
                    (name, capacity, description, location_type, self.current_user['id'])
                )
                
                self.db.log_action(self.current_user['id'], 'add_warehouse',
                                  f'إضافة مستودع جديد {name}')
                
                dialog.open = False
                self.page.update()
                self.show_snack_bar("تم إضافة المستودع بنجاح", COLORS['success'])
                self.load_warehouses()
                
            except sqlite3.IntegrityError:
                self.show_snack_bar("اسم المستودع موجود مسبقاً", COLORS['danger'])
        
        dialog = ft.AlertDialog(
            title=ft.Text("إضافة مستودع جديد", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    name_field,
                    capacity_field,
                    desc_field,
                    type_dropdown,
                ], spacing=15),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_warehouse, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def edit_warehouse(self, warehouse_id, name):
        """تعديل بيانات المستودع"""
        if not self.check_permission('can_edit_warehouse'):
            self.show_snack_bar("غير مصرح لك بتعديل المستودعات", COLORS['danger'])
            return
        
        result = self.db.execute_query(
            "SELECT capacity, description, location_type FROM warehouses WHERE id = ?",
            (warehouse_id,)
        )
        
        if not result:
            self.show_snack_bar("المستودع غير موجود", COLORS['danger'])
            return
        
        capacity, description, location_type = result[0]
        
        # حقول الإدخال
        name_display = ft.TextField(
            label="اسم المستودع",
            width=300,
            value=name,
            read_only=True,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT
        )
        
        capacity_field = ft.TextField(
            label="السعة",
            width=300,
            value=str(capacity),
            keyboard_type=ft.KeyboardType.NUMBER,
            text_align=ft.TextAlign.RIGHT
        )
        
        desc_field = ft.TextField(
            label="الوصف",
            width=300,
            value=description or "",
            text_align=ft.TextAlign.RIGHT
        )
        
        type_dropdown = ft.Dropdown(
            label="نوع المستودع",
            width=300,
            options=[
                ft.dropdown.Option("main", "رئيسي"),
                ft.dropdown.Option("external", "خارجي"),
                ft.dropdown.Option("north", "شمالي"),
                ft.dropdown.Option("south", "جنوبي"),
                ft.dropdown.Option("other", "آخر"),
            ],
            value=location_type or "other"
        )
        
        def save_edit(e):
            new_capacity_text = capacity_field.value.strip()
            new_description = desc_field.value or ""
            new_location_type = type_dropdown.value
            
            try:
                new_capacity = int(new_capacity_text) if new_capacity_text else capacity
            except ValueError:
                new_capacity = capacity
            
            self.db.execute_query(
                "UPDATE warehouses SET capacity = ?, description = ?, location_type = ? WHERE id = ?",
                (new_capacity, new_description, new_location_type, warehouse_id)
            )
            
            self.db.log_action(self.current_user['id'], 'edit_warehouse',
                              f'تعديل المستودع {name}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تحديث بيانات المستودع بنجاح", COLORS['success'])
            self.load_warehouses()
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تعديل المستودع: {name}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    name_display,
                    capacity_field,
                    desc_field,
                    type_dropdown,
                ], spacing=15),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_edit, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def delete_warehouse(self, warehouse_id, name):
        """حذف مستودع"""
        if not self.check_permission('can_delete_warehouse'):
            self.show_snack_bar("غير مصرح لك بحذف المستودعات", COLORS['danger'])
            return
        
        result = self.db.execute_query(
            "SELECT COUNT(*) FROM carts WHERE current_warehouse_id = ?",
            (warehouse_id,)
        )
        count = result[0][0] if result else 0
        
        if count > 0:
            self.show_snack_bar(f"لا يمكن حذف المستودع لأنه يحتوي على {count} عربة. قم بنقلها أولاً.", COLORS['danger'])
            return
        
        def confirm_delete(e):
            self.db.execute_query(
                "UPDATE warehouses SET is_active = 0 WHERE id = ?",
                (warehouse_id,)
            )
            
            self.db.log_action(self.current_user['id'], 'delete_warehouse',
                              f'حذف المستودع {name}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم حذف المستودع بنجاح", COLORS['success'])
            self.load_warehouses()
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد الحذف"),
            content=ft.Text(f"هل أنت متأكد من حذف المستودع '{name}'؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_delete),
                ft.TextButton("لا", on_click=cancel_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ================================ التقارير ================================
    def show_reports(self):
        """عرض صفحة التقارير"""
        if not self.check_permission('can_view_reports'):
            self.show_snack_bar("غير مصرح لك بعرض التقارير", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("التقارير والتحليلات", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== خيارات التقرير =====
        options_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("خيارات التقرير", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.ResponsiveRow([
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 4},
                        content=ft.Dropdown(
                            label="نوع التقرير",
                            options=[
                                ft.dropdown.Option("تقرير حالة العربات"),
                                ft.dropdown.Option("تقرير حركة العربات"),
                                ft.dropdown.Option("تقرير الصيانة"),
                                ft.dropdown.Option("تقرير المستودعات"),
                                ft.dropdown.Option("تقرير شامل"),
                            ],
                            value="تقرير حالة العربات",
                            on_change=self.update_report_preview,
                            ref=ft.Ref[ft.Dropdown]()
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 6, "lg": 4},
                        content=ft.Dropdown(
                            label="الفترة",
                            options=[
                                ft.dropdown.Option("اليوم"),
                                ft.dropdown.Option("آخر 7 أيام"),
                                ft.dropdown.Option("آخر 30 يوم"),
                                ft.dropdown.Option("آخر سنة"),
                                ft.dropdown.Option("كل الفترات"),
                            ],
                            value="كل الفترات",
                            on_change=self.update_report_preview,
                            ref=ft.Ref[ft.Dropdown]()
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 12, "lg": 4},
                        content=ft.Row([
                            ft.ElevatedButton(
                                text="معاينة التقرير",
                                icon=ft.icons.PREVIEW,
                                bgcolor=COLORS['primary'],
                                color=COLORS['white'],
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.update_report_preview
                            ),
                            ft.ElevatedButton(
                                text="تصدير Excel",
                                icon=ft.icons.TABLE_CHART,
                                bgcolor=COLORS['success'],
                                color=COLORS['white'],
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.export_to_excel,
                                visible=self.check_permission('can_export_reports') and EXCEL_AVAILABLE
                            ),
                            ft.ElevatedButton(
                                text="تصدير PDF",
                                icon=ft.icons.PICTURE_AS_PDF,
                                bgcolor=COLORS['danger'],
                                color=COLORS['white'],
                                style=ft.ButtonStyle(
                                    shape=ft.RoundedRectangleBorder(radius=8),
                                ),
                                on_click=self.export_to_pdf,
                                visible=self.check_permission('can_export_reports') and FPDF_AVAILABLE
                            ),
                        ], alignment=ft.MainAxisAlignment.END)
                    ),
                ])
            ])
        )
        
        self.report_type_dropdown = options_card.content.controls[2].controls[0].content
        self.period_dropdown = options_card.content.controls[2].controls[1].content
        
        self.content_column.controls.append(options_card)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== معاينة التقرير =====
        preview_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            expand=True,
            content=ft.Column([
                ft.Text("معاينة التقرير", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.DataTable(
                        columns=[],
                        rows=[],
                        horizontal_margin=10,
                        column_spacing=20,
                        heading_row_color=COLORS['light'],
                        heading_row_height=40,
                        data_row_max_height=40,
                        expand=True,
                        ref=ft.Ref[ft.DataTable]()
                    ),
                    expand=True,
                    scroll=ft.ScrollMode.AUTO
                )
            ], expand=True)
        )
        
        self.preview_table = preview_card.content.controls[2].content
        self.content_column.controls.append(preview_card)
        
        self.update_report_preview(None)
        self.page.update()
    
    def update_report_preview(self, e):
        """تحديث معاينة التقرير"""
        if not self.preview_table:
            return
        
        report_type = self.report_type_dropdown.value if self.report_type_dropdown else "تقرير حالة العربات"
        period = self.period_dropdown.value if self.period_dropdown else "كل الفترات"
        
        try:
            if report_type == "تقرير حالة العربات":
                self.preview_cart_status_report()
            elif report_type == "تقرير حركة العربات":
                self.preview_movement_report(period)
            elif report_type == "تقرير الصيانة":
                self.preview_maintenance_report(period)
            elif report_type == "تقرير المستودعات":
                self.preview_warehouse_report()
            elif report_type == "تقرير شامل":
                self.preview_summary_report()
        except Exception as ex:
            self.show_snack_bar(f"حدث خطأ أثناء إنشاء المعاينة: {str(ex)}", COLORS['danger'])
    
    def preview_cart_status_report(self):
        """معاينة تقرير حالة العربات"""
        self.preview_table.columns = [
            ft.DataColumn(ft.Text("الحالة", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("العدد", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("النسبة المئوية", size=14, weight=ft.FontWeight.BOLD)),
        ]
        self.preview_table.rows.clear()
        
        data = self.db.execute_query("""
            SELECT 
                status,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM carts), 2) as percentage
            FROM carts
            GROUP BY status
            UNION
            SELECT 'الإجمالي', COUNT(*), 100.0 FROM carts
        """)
        
        for row in data:
            status, count, percentage = row
            status_text = CART_STATUS.get(status, status) if status != 'الإجمالي' else status
            
            self.preview_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(status_text, size=13)),
                        ft.DataCell(ft.Text(str(count), size=13)),
                        ft.DataCell(ft.Text(f"{percentage}%", size=13)),
                    ]
                )
            )
        
        self.page.update()
    
    def preview_movement_report(self, period):
        """معاينة تقرير حركة العربات"""
        self.preview_table.columns = [
            ft.DataColumn(ft.Text("التاريخ", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("عدد الحركات", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("عربات مختلفة", size=14, weight=ft.FontWeight.BOLD)),
        ]
        self.preview_table.rows.clear()
        
        limit = ""
        if period == "اليوم":
            limit = "AND DATE(timestamp) = DATE('now')"
        elif period == "آخر 7 أيام":
            limit = "AND DATE(timestamp) >= DATE('now', '-7 days')"
        elif period == "آخر 30 يوم":
            limit = "AND DATE(timestamp) >= DATE('now', '-30 days')"
        elif period == "آخر سنة":
            limit = "AND DATE(timestamp) >= DATE('now', '-1 year')"
        
        query = f"""
            SELECT 
                DATE(timestamp) as date,
                COUNT(*) as movements,
                COUNT(DISTINCT cart_id) as carts_moved
            FROM movements
            WHERE 1=1 {limit}
            GROUP BY DATE(timestamp)
            ORDER BY date DESC
            LIMIT 10
        """
        
        data = self.db.execute_query(query)
        
        for row in data:
            self.preview_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(row[0] or "", size=13)),
                        ft.DataCell(ft.Text(str(row[1]), size=13)),
                        ft.DataCell(ft.Text(str(row[2]), size=13)),
                    ]
                )
            )
        
        self.page.update()
    
    def preview_maintenance_report(self, period):
        """معاينة تقرير الصيانة"""
        self.preview_table.columns = [
            ft.DataColumn(ft.Text("حالة الصيانة", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("العدد", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("التكلفة الإجمالية", size=14, weight=ft.FontWeight.BOLD)),
        ]
        self.preview_table.rows.clear()
        
        limit = ""
        if period == "اليوم":
            limit = "AND DATE(entry_date) = DATE('now')"
        elif period == "آخر 7 أيام":
            limit = "AND DATE(entry_date) >= DATE('now', '-7 days')"
        elif period == "آخر 30 يوم":
            limit = "AND DATE(entry_date) >= DATE('now', '-30 days')"
        elif period == "آخر سنة":
            limit = "AND DATE(entry_date) >= DATE('now', '-1 year')"
        
        query = f"""
            SELECT 
                status,
                COUNT(*) as count,
                SUM(cost) as total_cost
            FROM maintenance_records
            WHERE 1=1 {limit}
            GROUP BY status
        """
        
        data = self.db.execute_query(query)
        
        for row in data:
            status, count, total_cost = row
            status_text = MAINTENANCE_STATUS.get(status, status)
            
            self.preview_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(status_text, size=13)),
                        ft.DataCell(ft.Text(str(count), size=13)),
                        ft.DataCell(ft.Text(f"{total_cost or 0:.0f} ر.س", size=13)),
                    ]
                )
            )
        
        self.page.update()
    
    def preview_warehouse_report(self):
        """معاينة تقرير المستودعات"""
        self.preview_table.columns = [
            ft.DataColumn(ft.Text("المستودع", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("السعة", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("العدد الحالي", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("نسبة الإشغال", size=14, weight=ft.FontWeight.BOLD)),
        ]
        self.preview_table.rows.clear()
        
        data = self.db.execute_query("""
            SELECT 
                name,
                capacity,
                current_count,
                ROUND(current_count * 100.0 / capacity, 2) as occupancy
            FROM warehouses
            WHERE is_active = 1
            ORDER BY occupancy DESC
        """)
        
        for row in data:
            name, capacity, current, occupancy = row
            
            self.preview_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(name, size=13)),
                        ft.DataCell(ft.Text(str(capacity), size=13)),
                        ft.DataCell(ft.Text(str(current), size=13)),
                        ft.DataCell(ft.Text(f"{occupancy}%", size=13)),
                    ]
                )
            )
        
        self.page.update()
    
    def preview_summary_report(self):
        """معاينة التقرير الشامل"""
        self.preview_table.columns = [
            ft.DataColumn(ft.Text("المؤشر", size=14, weight=ft.FontWeight.BOLD)),
            ft.DataColumn(ft.Text("القيمة", size=14, weight=ft.FontWeight.BOLD)),
        ]
        self.preview_table.rows.clear()
        
        total_carts = self.db.execute_query("SELECT COUNT(*) FROM carts")[0][0] or 0
        sound_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'sound'")[0][0] or 0
        maintenance_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'needs_maintenance'")[0][0] or 0
        damaged_carts = self.db.execute_query("SELECT COUNT(*) FROM carts WHERE status = 'damaged'")[0][0] or 0
        total_warehouses = self.db.execute_query("SELECT COUNT(*) FROM warehouses WHERE is_active = 1")[0][0] or 0
        total_movements = self.db.execute_query("SELECT COUNT(*) FROM movements")[0][0] or 0
        total_maintenance = self.db.execute_query("SELECT COUNT(*) FROM maintenance_records")[0][0] or 0
        total_cost = self.db.execute_query("SELECT SUM(cost) FROM maintenance_records WHERE status = 'completed'")[0][0] or 0
        total_users = self.db.execute_query("SELECT COUNT(*) FROM users WHERE is_active = 1")[0][0] or 0
        
        summary_data = [
            ("إجمالي العربات", f"{total_carts} عربة"),
            ("عربات سليمة", f"{sound_carts} عربة ({sound_carts/total_carts*100:.1f}%)" if total_carts > 0 else "0"),
            ("تحتاج صيانة", f"{maintenance_carts} عربة ({maintenance_carts/total_carts*100:.1f}%)" if total_carts > 0 else "0"),
            ("عربات تالفة", f"{damaged_carts} عربة ({damaged_carts/total_carts*100:.1f}%)" if total_carts > 0 else "0"),
            ("عدد المستودعات", f"{total_warehouses} مستودع"),
            ("إجمالي الحركات", f"{total_movements} حركة"),
            ("عمليات الصيانة", f"{total_maintenance} عملية"),
            ("تكاليف الصيانة", f"{total_cost:.0f} ر.س"),
            ("المستخدمين النشطين", f"{total_users} مستخدم"),
            ("اسم المستخدم", self.current_user['username']),
            ("الدور", "مدير" if self.current_user['role'] == 'admin' else "مشغل"),
            ("تاريخ التقرير", datetime.now().strftime('%Y-%m-%d %H:%M'))
        ]
        
        for indicator, value in summary_data:
            self.preview_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(indicator, size=13)),
                        ft.DataCell(ft.Text(value, size=13)),
                    ]
                )
            )
        
        self.page.update()
    
    def export_to_excel(self, e):
        """تصدير التقرير إلى Excel"""
        if not self.check_permission('can_export_reports'):
            self.show_snack_bar("غير مصرح لك بتصدير التقارير", COLORS['danger'])
            return
        
        if not EXCEL_AVAILABLE:
            self.show_snack_bar("مكتبة openpyxl غير مثبتة", COLORS['danger'])
            return
        
        try:
            from tkinter import filedialog, Tk
            
            root = Tk()
            root.withdraw()
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"تقرير_{self.report_type_dropdown.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            root.destroy()
            
            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = "تقرير"
                
                ws['A1'] = f"تقرير: {self.report_type_dropdown.value}"
                ws['A2'] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                ws['A3'] = f"المستخدم: {self.current_user['username']}"
                
                # إضافة الرؤوس
                headers = [col.label.value for col in self.preview_table.columns]
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=5, column=col_num, value=header)
                
                # إضافة البيانات
                row_num = 6
                for row in self.preview_table.rows:
                    for col_num, cell in enumerate(row.cells, 1):
                        if isinstance(cell.content, ft.Text):
                            ws.cell(row=row_num, column=col_num, value=cell.content.value)
                        elif isinstance(cell.content, ft.Container):
                            if isinstance(cell.content.content, ft.Text):
                                ws.cell(row=row_num, column=col_num, value=cell.content.content.value)
                    row_num += 1
                
                wb.save(filename)
                
                self.db.log_action(self.current_user['id'], 'export_excel',
                                  f'تصدير تقرير {self.report_type_dropdown.value} إلى Excel')
                
                self.show_snack_bar(f"تم حفظ التقرير بنجاح", COLORS['success'])
                
        except Exception as ex:
            self.show_snack_bar(f"حدث خطأ أثناء حفظ الملف: {str(ex)}", COLORS['danger'])
    
    def export_to_pdf(self, e):
        """تصدير التقرير إلى PDF"""
        if not self.check_permission('can_export_reports'):
            self.show_snack_bar("غير مصرح لك بتصدير التقارير", COLORS['danger'])
            return
        
        if not FPDF_AVAILABLE:
            self.show_snack_bar("مكتبة fpdf غير مثبتة", COLORS['danger'])
            return
        
        try:
            from tkinter import filedialog, Tk
            
            root = Tk()
            root.withdraw()
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialfile=f"تقرير_{self.report_type_dropdown.value}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            root.destroy()
            
            if filename:
                pdf = FPDF()
                pdf.add_page()
                
                # محاولة استخدام خط عربي
                try:
                    font_path = self.find_arabic_font()
                    if font_path:
                        pdf.add_font("Arabic", "", font_path, uni=True)
                        pdf.set_font("Arabic", "", 16)
                    else:
                        pdf.set_font("Arial", "", 16)
                except:
                    pdf.set_font("Arial", "", 16)
                
                # عنوان التقرير
                pdf.cell(200, 10, txt=self.report_type_dropdown.value, ln=1, align='C')
                
                pdf.set_font_size(12)
                pdf.cell(200, 10, txt=f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=1, align='C')
                pdf.cell(200, 10, txt=f"المستخدم: {self.current_user['username']}", ln=1, align='C')
                pdf.ln(10)
                
                # جدول البيانات
                col_width = pdf.w / (len(self.preview_table.columns) + 1)
                pdf.set_font_size(10)
                
                # رؤوس الأعمدة
                for col in self.preview_table.columns:
                    pdf.cell(col_width, 10, col.label.value, border=1, align='C')
                pdf.ln()
                
                # بيانات الجدول
                for row in self.preview_table.rows:
                    for cell in row.cells:
                        if isinstance(cell.content, ft.Text):
                            pdf.cell(col_width, 10, cell.content.value, border=1, align='C')
                        elif isinstance(cell.content, ft.Container):
                            if isinstance(cell.content.content, ft.Text):
                                pdf.cell(col_width, 10, cell.content.content.value, border=1, align='C')
                    pdf.ln()
                
                pdf.output(filename)
                
                self.db.log_action(self.current_user['id'], 'export_pdf',
                                  f'تصدير تقرير {self.report_type_dropdown.value} إلى PDF')
                
                self.show_snack_bar(f"تم حفظ التقرير بنجاح", COLORS['success'])
                
        except Exception as ex:
            self.show_snack_bar(f"حدث خطأ أثناء إنشاء ملف PDF: {str(ex)}", COLORS['danger'])
    
    def find_arabic_font(self):
        """البحث عن خط عربي في النظام"""
        possible_paths = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "DejaVuSans.ttf"),
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "C:\\Windows\\Fonts\\arial.ttf",
            "C:\\Windows\\Fonts\\tahoma.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                return path
        return None
    
    # ================================ إدارة المستخدمين ================================
    def show_user_management(self):
        """عرض صفحة إدارة المستخدمين"""
        if self.current_user['role'] != 'admin':
            self.show_snack_bar("غير مصرح لك بالوصول إلى هذه الصفحة", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        title_row = ft.Row([
            ft.Text("إدارة المستخدمين والصلاحيات", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
            ft.Row([
                ft.TextField(
                    hint_text="بحث...",
                    width=250,
                    height=40,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    prefix=ft.Icon(ft.icons.SEARCH),
                    on_change=self.filter_users,
                    ref=ft.Ref[ft.TextField]()
                ),
                ft.ElevatedButton(
                    text="إضافة مستخدم جديد",
                    icon=ft.icons.PERSON_ADD,
                    bgcolor=COLORS['success'],
                    color=COLORS['white'],
                    style=ft.ButtonStyle(
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=self.show_add_user_dialog
                ),
                ft.ElevatedButton(
                    text="تغيير كلمة مرور المدير",
                    icon=ft.icons.LOCK_RESET,
                    bgcolor=COLORS['warning'],
                    color=COLORS['white'],
                    style=ft.ButtonStyle(
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=self.show_change_admin_password
                ),
            ])
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
        
        self.content_column.controls.append(title_row)
        self.content_column.controls.append(ft.Container(height=20))
        
        # تخزين مرجع حقل البحث
        self.user_search_field = title_row.controls[1].controls[0]
        
        # جدول المستخدمين
        self.user_table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("المعرف", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("اسم المستخدم", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الاسم الكامل", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الدور", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الحالة", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("آخر تسجيل", size=14, weight=ft.FontWeight.BOLD)),
                ft.DataColumn(ft.Text("الإجراءات", size=14, weight=ft.FontWeight.BOLD)),
            ],
            rows=[],
            horizontal_margin=10,
            column_spacing=20,
            heading_row_color=COLORS['light'],
            heading_row_height=50,
            data_row_max_height=60,
            expand=True
        )
        
        # حاوية الجدول مع التمرير
        table_container = ft.Container(
            content=ft.Column([
                self.user_table
            ], scroll=ft.ScrollMode.AUTO),
            expand=True,
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=15
        )
        
        self.content_column.controls.append(table_container)
        self.load_users()
        self.page.update()
    
    def load_users(self):
        """تحميل قائمة المستخدمين"""
        if not self.user_table:
            return
        
        self.user_table.rows.clear()
        
        users = self.db.execute_query("""
            SELECT id, username, full_name, role, is_active, last_login 
            FROM users 
            ORDER BY id
        """)
        
        for user in users:
            uid, username, full_name, role, is_active, last_login = user
            role_text = "مدير" if role == 'admin' else "مشغل"
            status_text = "نشط" if is_active else "غير نشط"
            status_color = COLORS['success'] if is_active else COLORS['danger']
            last_login_text = last_login[:16] if last_login else "لم يسجل دخول"
            
            # أزرار الإجراءات
            actions_row = ft.Row(spacing=5)
            
            actions_row.controls.append(
                ft.IconButton(
                    icon=ft.icons.EDIT,
                    icon_size=18,
                    icon_color=COLORS['primary'],
                    tooltip="تعديل",
                    on_click=lambda e, uid=uid, un=username: self.edit_user(uid, un)
                )
            )
            
            actions_row.controls.append(
                ft.IconButton(
                    icon=ft.icons.SECURITY,
                    icon_size=18,
                    icon_color=COLORS['purple'],
                    tooltip="صلاحيات",
                    on_click=lambda e, uid=uid, un=username: self.manage_user_permissions(uid, un)
                )
            )
            
            actions_row.controls.append(
                ft.IconButton(
                    icon=ft.icons.LOCK_RESET,
                    icon_size=18,
                    icon_color=COLORS['warning'],
                    tooltip="تغيير كلمة المرور",
                    on_click=lambda e, uid=uid, un=username: self.change_password(uid, un)
                )
            )
            
            if username != DEFAULT_USER:
                actions_row.controls.append(
                    ft.IconButton(
                        icon=ft.icons.DELETE,
                        icon_size=18,
                        icon_color=COLORS['danger'],
                        tooltip="حذف",
                        on_click=lambda e, uid=uid, un=username: self.delete_user(uid, un)
                    )
                )
                
                status_icon = ft.IconButton(
                    icon=ft.icons.CANCEL if is_active else ft.icons.CHECK_CIRCLE,
                    icon_size=18,
                    icon_color=COLORS['danger'] if is_active else COLORS['success'],
                    tooltip="تعطيل" if is_active else "تفعيل",
                    on_click=lambda e, uid=uid, un=username, act=not is_active: 
                        self.toggle_user_status(uid, un, act)
                )
                actions_row.controls.append(status_icon)
            
            self.user_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(uid), size=13)),
                        ft.DataCell(ft.Text(username, size=13)),
                        ft.DataCell(ft.Text(full_name or "", size=13)),
                        ft.DataCell(ft.Text(role_text, size=13)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(status_text, size=12, color=COLORS['white']),
                            bgcolor=status_color,
                            padding=ft.padding.symmetric(horizontal=8, vertical=4),
                            border_radius=4
                        )),
                        ft.DataCell(ft.Text(last_login_text, size=13)),
                        ft.DataCell(actions_row),
                    ]
                )
            )
        
        self.page.update()
    
    def filter_users(self, e):
        """فلترة المستخدمين حسب البحث"""
        if not self.user_table:
            return
        
        search_text = e.control.value.strip().lower() if e.control.value else ""
        
        for row in self.user_table.rows[:]:
            username_cell = row.cells[1].content
            fullname_cell = row.cells[2].content
            
            match = False
            if isinstance(username_cell, ft.Text):
                if search_text in username_cell.value.lower():
                    match = True
            if isinstance(fullname_cell, ft.Text):
                if search_text in fullname_cell.value.lower():
                    match = True
            
            if search_text and not match:
                self.user_table.rows.remove(row)
        
        self.page.update()
    
    def show_add_user_dialog(self, e):
        """عرض نافذة إضافة مستخدم جديد"""
        # حقول الإدخال
        username_field = ft.TextField(
            label="اسم المستخدم",
            width=300,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT,
            autofocus=True
        )
        
        password_field = ft.TextField(
            label="كلمة المرور",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT
        )
        
        confirm_field = ft.TextField(
            label="تأكيد كلمة المرور",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT
        )
        
        fullname_field = ft.TextField(
            label="الاسم الكامل",
            width=300,
            text_align=ft.TextAlign.RIGHT
        )
        
        role_dropdown = ft.Dropdown(
            label="الدور",
            width=300,
            options=[
                ft.dropdown.Option("مشغل"),
                ft.dropdown.Option("مدير"),
            ],
            value="مشغل"
        )
        
        def save_user(e):
            username = username_field.value.strip() if username_field.value else ""
            password = password_field.value.strip() if password_field.value else ""
            confirm = confirm_field.value.strip() if confirm_field.value else ""
            fullname = fullname_field.value or ""
            role_text = role_dropdown.value
            
            if not username or not password:
                self.show_snack_bar("الرجاء إدخال اسم المستخدم وكلمة المرور", COLORS['danger'])
                return
            
            if password != confirm:
                self.show_snack_bar("كلمة المرور غير متطابقة", COLORS['danger'])
                return
            
            role = "admin" if role_text == "مدير" else "operator"
            
            try:
                user_id = self.db.execute_insert(
                    """INSERT INTO users (username, password, full_name, role, is_active, created_by) 
                       VALUES (?, ?, ?, ?, 1, ?)""",
                    (username, password, fullname, role, self.current_user['id'])
                )
                
                permissions = DEFAULT_PERMISSIONS.copy()
                if role == 'admin':
                    for key in permissions:
                        permissions[key] = 1
                
                self.db.update_user_permissions(user_id, permissions)
                self.db.log_action(self.current_user['id'], 'add_user',
                                  f'إضافة مستخدم جديد {username}')
                
                dialog.open = False
                self.page.update()
                self.show_snack_bar("تم إضافة المستخدم بنجاح", COLORS['success'])
                self.load_users()
                
            except sqlite3.IntegrityError:
                self.show_snack_bar("اسم المستخدم موجود مسبقاً", COLORS['danger'])
        
        dialog = ft.AlertDialog(
            title=ft.Text("إضافة مستخدم جديد", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    username_field,
                    password_field,
                    confirm_field,
                    fullname_field,
                    role_dropdown,
                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_user, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def edit_user(self, user_id, username):
        """تعديل بيانات المستخدم"""
        result = self.db.execute_query(
            "SELECT full_name, role FROM users WHERE id = ?",
            (user_id,)
        )
        
        if not result:
            self.show_snack_bar("المستخدم غير موجود", COLORS['danger'])
            return
        
        fullname, role = result[0]
        
        # حقول الإدخال
        username_display = ft.TextField(
            label="اسم المستخدم",
            width=300,
            value=username,
            read_only=True,
            border_radius=8,
            text_align=ft.TextAlign.RIGHT
        )
        
        fullname_field = ft.TextField(
            label="الاسم الكامل",
            width=300,
            value=fullname or "",
            text_align=ft.TextAlign.RIGHT
        )
        
        role_dropdown = ft.Dropdown(
            label="الدور",
            width=300,
            options=[
                ft.dropdown.Option("مشغل"),
                ft.dropdown.Option("مدير"),
            ],
            value="مدير" if role == 'admin' else "مشغل"
        )
        
        def save_edit(e):
            new_fullname = fullname_field.value or ""
            new_role_text = role_dropdown.value
            new_role = "admin" if new_role_text == "مدير" else "operator"
            
            self.db.execute_query(
                "UPDATE users SET full_name = ?, role = ? WHERE id = ?",
                (new_fullname, new_role, user_id)
            )
            
            if new_role == 'admin':
                permissions = DEFAULT_PERMISSIONS.copy()
                for key in permissions:
                    permissions[key] = 1
                self.db.update_user_permissions(user_id, permissions)
            
            self.db.log_action(self.current_user['id'], 'edit_user',
                              f'تعديل بيانات المستخدم {username}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تحديث بيانات المستخدم بنجاح", COLORS['success'])
            self.load_users()
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تعديل بيانات المستخدم: {username}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    username_display,
                    fullname_field,
                    role_dropdown,
                ], spacing=15),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_edit, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def manage_user_permissions(self, user_id, username):
        """إدارة صلاحيات المستخدم"""
        permissions = self.db.get_user_permissions(user_id)
        
        # مجموعات الصلاحيات
        groups = [
            ("لوحة التحكم", ['can_view_dashboard']),
            ("إدارة العربات", ['can_manage_carts', 'can_add_cart', 'can_edit_cart', 'can_delete_cart']),
            ("حركة العربات", ['can_move_cart', 'can_view_movements']),
            ("الصيانة", ['can_manage_maintenance', 'can_complete_maintenance']),
            ("المستودعات", ['can_view_warehouses', 'can_add_warehouse', 'can_edit_warehouse', 'can_delete_warehouse']),
            ("التقارير", ['can_view_reports', 'can_export_reports']),
            ("إدارة النظام", ['can_manage_users', 'can_manage_backup']),
            ("إعدادات المستخدم", ['can_change_own_password'])
        ]
        
        permission_labels = {
            'can_view_dashboard': 'عرض لوحة التحكم',
            'can_manage_carts': 'إدارة العربات',
            'can_add_cart': 'إضافة عربة جديدة',
            'can_edit_cart': 'تعديل العربات',
            'can_delete_cart': 'حذف العربات',
            'can_move_cart': 'نقل العربات',
            'can_view_movements': 'عرض سجل الحركات',
            'can_manage_maintenance': 'إدارة الصيانة',
            'can_complete_maintenance': 'إتمام الصيانة',
            'can_view_warehouses': 'عرض المستودعات',
            'can_add_warehouse': 'إضافة مستودع',
            'can_edit_warehouse': 'تعديل المستودعات',
            'can_delete_warehouse': 'حذف المستودعات',
            'can_view_reports': 'عرض التقارير',
            'can_export_reports': 'تصدير التقارير',
            'can_manage_users': 'إدارة المستخدمين',
            'can_manage_backup': 'إدارة النسخ الاحتياطي',
            'can_change_own_password': 'تغيير كلمة المرور الشخصية'
        }
        
        permission_vars = {}
        permission_controls = []
        
        for group_name, perm_list in groups:
            group_controls = []
            for perm in perm_list:
                if perm in permission_labels:
                    var = ft.Checkbox(
                        label=permission_labels[perm],
                        value=permissions.get(perm, 0) == 1,
                        fill_color=COLORS['primary']
                    )
                    permission_vars[perm] = var
                    group_controls.append(var)
            
            if group_controls:
                permission_controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text(group_name, size=16, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                            ft.Column(group_controls, spacing=5),
                            ft.Divider(height=1, color=COLORS['light']),
                        ]),
                        padding=10
                    )
                )
        
        def select_all(e):
            for var in permission_vars.values():
                var.value = True
            self.page.update()
        
        def deselect_all(e):
            for var in permission_vars.values():
                var.value = False
            self.page.update()
        
        def save_permissions(e):
            new_permissions = {}
            for key, var in permission_vars.items():
                new_permissions[key] = 1 if var.value else 0
            
            self.db.update_user_permissions(user_id, new_permissions)
            self.db.log_action(self.current_user['id'], 'edit_permissions',
                              f'تعديل صلاحيات المستخدم {username}')
            
            if user_id == self.current_user['id']:
                self.current_permissions = self.db.get_user_permissions(user_id)
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar(f"تم تحديث صلاحيات المستخدم {username} بنجاح", COLORS['success'])
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"صلاحيات المستخدم: {username}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=600,
                height=500,
                content=ft.Column(
                    permission_controls,
                    scroll=ft.ScrollMode.AUTO,
                    spacing=10
                ),
                padding=10
            ),
            actions=[
                ft.TextButton("تحديد الكل", on_click=select_all),
                ft.TextButton("إلغاء الكل", on_click=deselect_all),
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_permissions, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def show_change_admin_password(self, e):
        """تغيير كلمة مرور المدير الرئيسي"""
        # حقول الإدخال
        current_pass_field = ft.TextField(
            label="كلمة المرور الحالية",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT,
            autofocus=True
        )
        
        new_pass_field = ft.TextField(
            label="كلمة المرور الجديدة",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT
        )
        
        confirm_pass_field = ft.TextField(
            label="تأكيد كلمة المرور",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT
        )
        
        def save_password(e):
            current_pass = current_pass_field.value.strip() if current_pass_field.value else ""
            new_pass = new_pass_field.value.strip() if new_pass_field.value else ""
            confirm_pass = confirm_pass_field.value.strip() if confirm_pass_field.value else ""
            
            result = self.db.execute_query(
                "SELECT id FROM users WHERE username = ? AND password = ?",
                (DEFAULT_USER, current_pass)
            )
            
            if not result:
                self.show_snack_bar("كلمة المرور الحالية غير صحيحة", COLORS['danger'])
                return
            
            if not new_pass:
                self.show_snack_bar("الرجاء إدخال كلمة المرور الجديدة", COLORS['danger'])
                return
            
            if new_pass != confirm_pass:
                self.show_snack_bar("كلمة المرور غير متطابقة", COLORS['danger'])
                return
            
            admin_id = self.db.execute_query(
                "SELECT id FROM users WHERE username = ?",
                (DEFAULT_USER,)
            )[0][0]
            
            self.db.execute_query(
                "UPDATE users SET password = ? WHERE id = ?",
                (new_pass, admin_id)
            )
            
            self.db.log_action(self.current_user['id'], 'change_admin_password',
                              'تغيير كلمة مرور المدير الرئيسي')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تغيير كلمة مرور المدير بنجاح", COLORS['success'])
        
        dialog = ft.AlertDialog(
            title=ft.Text("تغيير كلمة مرور المدير", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    current_pass_field,
                    new_pass_field,
                    confirm_pass_field,
                ], spacing=15),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_password, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def change_password(self, user_id, username):
        """تغيير كلمة مرور مستخدم"""
        # حقول الإدخال
        new_pass_field = ft.TextField(
            label="كلمة المرور الجديدة",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT,
            autofocus=True
        )
        
        confirm_pass_field = ft.TextField(
            label="تأكيد كلمة المرور",
            width=300,
            password=True,
            can_reveal_password=True,
            text_align=ft.TextAlign.RIGHT
        )
        
        def save_password(e):
            new_pass = new_pass_field.value.strip() if new_pass_field.value else ""
            confirm_pass = confirm_pass_field.value.strip() if confirm_pass_field.value else ""
            
            if not new_pass:
                self.show_snack_bar("الرجاء إدخال كلمة المرور الجديدة", COLORS['danger'])
                return
            
            if new_pass != confirm_pass:
                self.show_snack_bar("كلمة المرور غير متطابقة", COLORS['danger'])
                return
            
            self.db.execute_query(
                "UPDATE users SET password = ? WHERE id = ?",
                (new_pass, user_id)
            )
            
            self.db.log_action(self.current_user['id'], 'change_password',
                              f'تغيير كلمة مرور المستخدم {username}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم تغيير كلمة المرور بنجاح", COLORS['success'])
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تغيير كلمة المرور - {username}", size=18, weight=ft.FontWeight.BOLD),
            content=ft.Container(
                width=350,
                content=ft.Column([
                    new_pass_field,
                    confirm_pass_field,
                ], spacing=15),
                padding=10
            ),
            actions=[
                ft.TextButton("إلغاء", on_click=lambda e: self.close_dialog(dialog)),
                ft.ElevatedButton("حفظ", on_click=save_password, bgcolor=COLORS['success'], color=COLORS['white']),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def toggle_user_status(self, user_id, username, activate):
        """تفعيل/تعطيل مستخدم"""
        status_text = "تفعيل" if activate else "تعطيل"
        
        def confirm_toggle(e):
            self.db.execute_query(
                "UPDATE users SET is_active = ? WHERE id = ?",
                (1 if activate else 0, user_id)
            )
            
            self.db.log_action(self.current_user['id'], 'toggle_user',
                              f'{status_text} المستخدم {username}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar(f"تم {status_text} المستخدم بنجاح", COLORS['success'])
            self.load_users()
        
        def cancel_toggle(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text(f"تأكيد {status_text} المستخدم"),
            content=ft.Text(f"هل أنت متأكد من {status_text} المستخدم '{username}'؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_toggle),
                ft.TextButton("لا", on_click=cancel_toggle),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    def delete_user(self, user_id, username):
        """حذف مستخدم"""
        if username == DEFAULT_USER:
            self.show_snack_bar("لا يمكن حذف المستخدم الرئيسي", COLORS['danger'])
            return
        
        def confirm_delete(e):
            self.db.execute_query("DELETE FROM users WHERE id = ?", (user_id,))
            self.db.log_action(self.current_user['id'], 'delete_user',
                              f'حذف المستخدم {username}')
            
            dialog.open = False
            self.page.update()
            self.show_snack_bar("تم حذف المستخدم بنجاح", COLORS['success'])
            self.load_users()
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("تأكيد الحذف"),
            content=ft.Text(f"هل أنت متأكد من حذف المستخدم '{username}'؟"),
            actions=[
                ft.TextButton("نعم", on_click=confirm_delete),
                ft.TextButton("لا", on_click=cancel_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.dialog = dialog
        dialog.open = True
        self.page.update()
    
    # ================================ إعدادات النظام ================================
    def show_system_settings(self):
        """عرض صفحة إعدادات النظام"""
        if self.current_user['role'] != 'admin':
            self.show_snack_bar("غير مصرح لك بالوصول إلى هذه الصفحة", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("إعدادات النظام", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== إعدادات التطبيق =====
        app_settings_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("إعدادات التطبيق", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text("اسم البرنامج:", size=14, weight=ft.FontWeight.BOLD),
                        ft.TextField(
                            value=self.db.get_app_setting('app_name', APP_NAME),
                            width=400,
                            border_radius=8,
                            text_align=ft.TextAlign.RIGHT,
                            ref=ft.Ref[ft.TextField]()
                        ),
                        ft.ElevatedButton(
                            text="حفظ اسم البرنامج",
                            icon=ft.icons.SAVE,
                            bgcolor=COLORS['primary'],
                            color=COLORS['white'],
                            style=ft.ButtonStyle(
                                shape=ft.RoundedRectangleBorder(radius=8),
                            ),
                            on_click=lambda e: self.save_app_name(e, app_name_field)
                        ),
                    ]),
                    padding=10
                ),
                
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text("اسم الجهة المشغلة:", size=14, weight=ft.FontWeight.BOLD),
                        ft.TextField(
                            value=self.db.get_app_setting('company_name', 
                                'الرئاسة العامة لشؤون المسجد الحرام والمسجد النبوي'),
                            width=400,
                            border_radius=8,
                            text_align=ft.TextAlign.RIGHT,
                            ref=ft.Ref[ft.TextField]()
                        ),
                        ft.ElevatedButton(
                            text="حفظ اسم الجهة",
                            icon=ft.icons.SAVE,
                            bgcolor=COLORS['primary'],
                            color=COLORS['white'],
                            style=ft.ButtonStyle(
                                shape=ft.RoundedRectangleBorder(radius=8),
                            ),
                            on_click=lambda e: self.save_company_name(e, company_name_field)
                        ),
                    ]),
                    padding=10
                )
            ])
        )
        
        # تخزين المراجع
        app_name_field = app_settings_card.content.controls[2].content.controls[1]
        company_name_field = app_settings_card.content.controls[4].content.controls[1]
        
        self.content_column.controls.append(app_settings_card)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== إعدادات MEGA =====
        mega_settings_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("إعدادات النسخ الاحتياطي السحابي (MEGA)", size=18, weight=ft.FontWeight.BOLD, 
                       color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(
                                name=ft.icons.CHECK_CIRCLE if MEGA_AVAILABLE else ft.icons.ERROR,
                                color=COLORS['success'] if MEGA_AVAILABLE else COLORS['danger']
                            ),
                            ft.Text(
                                "✓ مكتبة MEGA مثبتة - جاهز للعمل" if MEGA_AVAILABLE 
                                else "✗ مكتبة MEGA غير مثبتة - يرجى تثبيتها: pip install mega.py",
                                size=14,
                                color=COLORS['success'] if MEGA_AVAILABLE else COLORS['danger']
                            ),
                        ]),
                    ]),
                    padding=10
                ),
                
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text("البريد الإلكتروني MEGA:", size=14, weight=ft.FontWeight.BOLD),
                        ft.TextField(
                            value=self.db.get_app_setting('mega_email', MEGA_EMAIL),
                            width=400,
                            border_radius=8,
                            text_align=ft.TextAlign.RIGHT,
                            ref=ft.Ref[ft.TextField]()
                        ),
                        ft.Text("كلمة مرور MEGA:", size=14, weight=ft.FontWeight.BOLD),
                        ft.TextField(
                            value=self.db.get_app_setting('mega_password', MEGA_PASSWORD),
                            width=400,
                            border_radius=8,
                            text_align=ft.TextAlign.RIGHT,
                            password=True,
                            can_reveal_password=True,
                            ref=ft.Ref[ft.TextField]()
                        ),
                        ft.ElevatedButton(
                            text="حفظ إعدادات MEGA",
                            icon=ft.icons.SAVE,
                            bgcolor=COLORS['primary'],
                            color=COLORS['white'],
                            style=ft.ButtonStyle(
                                shape=ft.RoundedRectangleBorder(radius=8),
                            ),
                            on_click=lambda e: self.save_mega_settings(e, mega_email_field, mega_pass_field)
                        ),
                    ]),
                    padding=10
                )
            ])
        )
        
        # تخزين المراجع
        mega_email_field = mega_settings_card.content.controls[4].content.controls[1]
        mega_pass_field = mega_settings_card.content.controls[4].content.controls[3]
        
        self.content_column.controls.append(mega_settings_card)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== معلومات النظام =====
        info_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("معلومات النظام", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text(f"إصدار النظام: 2.0.0", size=14),
                        ft.Text(f"تاريخ الإصدار: 2025-02-12", size=14),
                        ft.Text(f"المطور: قسم تقنية المعلومات", size=14),
                        ft.Text(f"آخر تحديث: {datetime.now().strftime('%Y-%m-%d %H:%M')}", size=14),
                        ft.Text(f"دعم MEGA: {'مفعل ✓' if MEGA_AVAILABLE else 'غير مفعل ✗'}", size=14),
                    ]),
                    padding=10
                )
            ])
        )
        
        self.content_column.controls.append(info_card)
        self.page.update()
    
    def save_app_name(self, e, field):
        """حفظ اسم البرنامج"""
        new_name = field.value.strip()
        if new_name:
            self.db.update_app_setting('app_name', new_name, self.current_user['id'])
            self.db.log_action(self.current_user['id'], 'update_settings',
                              f'تحديث اسم البرنامج إلى: {new_name}')
            self.page.title = new_name
            self.show_snack_bar("تم تحديث اسم البرنامج بنجاح", COLORS['success'])
    
    def save_company_name(self, e, field):
        """حفظ اسم الجهة"""
        new_name = field.value.strip()
        if new_name:
            self.db.update_app_setting('company_name', new_name, self.current_user['id'])
            self.db.log_action(self.current_user['id'], 'update_settings',
                              f'تحديث اسم الجهة إلى: {new_name}')
            self.show_snack_bar("تم تحديث اسم الجهة بنجاح", COLORS['success'])
    
    def save_mega_settings(self, e, email_field, pass_field):
        """حفظ إعدادات MEGA"""
        new_email = email_field.value.strip()
        new_pass = pass_field.value.strip()
        
        if new_email:
            self.db.update_app_setting('mega_email', new_email, self.current_user['id'])
        if new_pass:
            self.db.update_app_setting('mega_password', new_pass, self.current_user['id'])
        
        self.db.log_action(self.current_user['id'], 'update_settings',
                          'تحديث إعدادات MEGA')
        
        self.show_snack_bar("تم تحديث إعدادات MEGA بنجاح", COLORS['success'])
    
    # ================================ تغيير كلمة المرور الشخصية ================================
    def show_change_password(self):
        """عرض صفحة تغيير كلمة المرور الشخصية"""
        if not self.check_permission('can_change_own_password'):
            self.show_snack_bar("غير مصرح لك بتغيير كلمة المرور", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("تغيير كلمة المرور الشخصية", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # بطاقة تغيير كلمة المرور
        change_pass_card = ft.Container(
            width=500,
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=30,
            content=ft.Column([
                ft.Text(f"المستخدم: {self.current_user['username']}", size=16, color=COLORS['gray']),
                ft.Divider(height=20, color=COLORS['light']),
                
                ft.TextField(
                    label="كلمة المرور الحالية",
                    password=True,
                    can_reveal_password=True,
                    width=400,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    ref=ft.Ref[ft.TextField]()
                ),
                
                ft.TextField(
                    label="كلمة المرور الجديدة",
                    password=True,
                    can_reveal_password=True,
                    width=400,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    ref=ft.Ref[ft.TextField]()
                ),
                
                ft.TextField(
                    label="تأكيد كلمة المرور",
                    password=True,
                    can_reveal_password=True,
                    width=400,
                    border_radius=8,
                    text_align=ft.TextAlign.RIGHT,
                    ref=ft.Ref[ft.TextField]()
                ),
                
                ft.Container(height=20),
                
                ft.ElevatedButton(
                    text="تغيير كلمة المرور",
                    icon=ft.icons.LOCK_RESET,
                    width=200,
                    height=45,
                    bgcolor=COLORS['success'],
                    color=COLORS['white'],
                    style=ft.ButtonStyle(
                        shape=ft.RoundedRectangleBorder(radius=8),
                    ),
                    on_click=lambda e: self.save_own_password(
                        e,
                        current_field,
                        new_field,
                        confirm_field
                    )
                )
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER)
        )
        
        # تخزين المراجع
        current_field = change_pass_card.content.controls[2]
        new_field = change_pass_card.content.controls[3]
        confirm_field = change_pass_card.content.controls[4]
        
        self.content_column.controls.append(
            ft.Container(
                content=change_pass_card,
                alignment=ft.alignment.center
            )
        )
        self.page.update()
    
    def save_own_password(self, e, current_field, new_field, confirm_field):
        """حفظ كلمة المرور الشخصية الجديدة"""
        current_pass = current_field.value.strip() if current_field.value else ""
        new_pass = new_field.value.strip() if new_field.value else ""
        confirm_pass = confirm_field.value.strip() if confirm_field.value else ""
        
        result = self.db.execute_query(
            "SELECT id FROM users WHERE id = ? AND password = ?",
            (self.current_user['id'], current_pass)
        )
        
        if not result:
            self.show_snack_bar("كلمة المرور الحالية غير صحيحة", COLORS['danger'])
            return
        
        if not new_pass:
            self.show_snack_bar("الرجاء إدخال كلمة المرور الجديدة", COLORS['danger'])
            return
        
        if new_pass != confirm_pass:
            self.show_snack_bar("كلمة المرور غير متطابقة", COLORS['danger'])
            return
        
        self.db.execute_query(
            "UPDATE users SET password = ? WHERE id = ?",
            (new_pass, self.current_user['id'])
        )
        
        self.db.log_action(self.current_user['id'], 'change_own_password',
                          'تغيير كلمة المرور الشخصية')
        
        self.show_snack_bar("تم تغيير كلمة المرور بنجاح", COLORS['success'])
    
    # ================================ النسخ الاحتياطي MEGA ================================
    def test_mega_connection(self):
        """اختبار الاتصال بـ MEGA"""
        if not MEGA_AVAILABLE:
            return False, "❌ مكتبة MEGA غير مثبتة. قم بتشغيل: pip install mega.py"
        
        mega_email = self.db.get_app_setting('mega_email', MEGA_EMAIL)
        mega_password = self.db.get_app_setting('mega_password', MEGA_PASSWORD)
        
        if not mega_email or not mega_password:
            return False, "❌ بيانات MEGA غير مكتملة. أضفها في ملف .env أو إعدادات النظام"
        
        try:
            mega = Mega()
            m = mega.login(mega_email, mega_password)
            account = m.get_user()
            email = account.get('email', mega_email)
            return True, f"✅ متصل بحساب: {email}"
        except Exception as e:
            error_msg = str(e)
            if "Invalid email or password" in error_msg:
                return False, "❌ البريد أو كلمة المرور غير صحيحين"
            elif "timeout" in error_msg.lower():
                return False, "❌ فشل الاتصال: تحقق من الإنترنت"
            else:
                return False, f"❌ خطأ: {error_msg[:50]}..."
    
    def show_backup(self):
        """عرض صفحة النسخ الاحتياطي"""
        if self.current_user['role'] != 'admin' or not self.check_permission('can_manage_backup'):
            self.show_snack_bar("غير مصرح لك بالوصول إلى هذه الصفحة", COLORS['danger'])
            return
        
        self.clear_content()
        
        # عنوان الصفحة
        self.content_column.controls.append(
            ft.Text("النسخ الاحتياطي", size=24, weight=ft.FontWeight.BOLD, color=COLORS['dark'])
        )
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== إنشاء نسخة احتياطية =====
        backup_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            content=ft.Column([
                ft.Text("إنشاء نسخة احتياطية", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.Row([
                    ft.ElevatedButton(
                        text="💾 نسخ احتياطي محلي",
                        icon=ft.icons.SAVE,
                        bgcolor=COLORS['primary'],
                        color=COLORS['white'],
                        style=ft.ButtonStyle(
                            shape=ft.RoundedRectangleBorder(radius=8),
                            padding=ft.padding.symmetric(horizontal=25, vertical=15)
                        ),
                        on_click=self.create_local_backup
                    ),
                    
                    ft.ElevatedButton(
                        text="☁️ نسخ احتياطي سحابي (MEGA)",
                        icon=ft.icons.CLOUD_UPLOAD,
                        bgcolor=COLORS['purple'] if MEGA_AVAILABLE else COLORS['gray'],
                        color=COLORS['white'],
                        style=ft.ButtonStyle(
                            shape=ft.RoundedRectangleBorder(radius=8),
                            padding=ft.padding.symmetric(horizontal=25, vertical=15)
                        ),
                        on_click=self.create_cloud_backup if MEGA_AVAILABLE else None,
                        disabled=not MEGA_AVAILABLE
                    ),
                ]),
                
                ft.Container(height=20),
                
                # شريط التقدم
                ft.ProgressBar(
                    width=400,
                    value=0,
                    bgcolor=COLORS['light'],
                    color=COLORS['primary'],
                    ref=ft.Ref[ft.ProgressBar]()
                ),
                
                ft.Container(height=10),
                
                # حالة النسخ
                ft.Text("", size=14, ref=ft.Ref[ft.Text]()),
            ])
        )
        
        # تخزين المراجع
        self.backup_progress = backup_card.content.controls[3].controls[2]
        self.backup_status = backup_card.content.controls[3].controls[4]
        
        self.content_column.controls.append(backup_card)
        self.content_column.controls.append(ft.Container(height=20))
        
        # ===== حالة MEGA =====
        if MEGA_AVAILABLE:
            mega_status = self.test_mega_connection()
            mega_status_card = ft.Container(
                bgcolor=COLORS['white'],
                border_radius=10,
                border=ft.border.all(1, COLORS['gray']),
                padding=20,
                content=ft.Row([
                    ft.Text("حالة MEGA:", size=14, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                    ft.Text(mega_status[1], size=14, color=COLORS['success'] if mega_status[0] else COLORS['danger']),
                ])
            )
            self.content_column.controls.append(mega_status_card)
            self.content_column.controls.append(ft.Container(height=20))
        
        # ===== سجل النسخ الاحتياطي =====
        history_card = ft.Container(
            bgcolor=COLORS['white'],
            border_radius=10,
            border=ft.border.all(1, COLORS['gray']),
            padding=20,
            expand=True,
            content=ft.Column([
                ft.Text("سجل النسخ الاحتياطي", size=18, weight=ft.FontWeight.BOLD, color=COLORS['dark']),
                ft.Divider(height=1, color=COLORS['light']),
                
                ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("التاريخ", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("اسم الملف", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("النوع", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("الحجم", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("رابط MEGA", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("الحالة", size=13, weight=ft.FontWeight.BOLD)),
                        ft.DataColumn(ft.Text("المستخدم", size=13, weight=ft.FontWeight.BOLD)),
                    ],
                    rows=[],
                    horizontal_margin=10,
                    column_spacing=15,
                    heading_row_color=COLORS['light'],
                    heading_row_height=40,
                    data_row_max_height=40,
                    expand=True,
                    ref=ft.Ref[ft.DataTable]()
                )
            ], expand=True)
        )
        
        self.backup_tree = history_card.content.controls[1]
        self.content_column.controls.append(history_card)
        
        self.load_backups()
        self.page.update()
    
    def create_local_backup(self, e):
        """إنشاء نسخة احتياطية محلية"""
        try:
            self.backup_progress.value = 0
            self.backup_status.value = "جاري إنشاء النسخة الاحتياطية..."
            self.backup_status.color = COLORS['primary']
            self.page.update()
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"backup_{timestamp}.db"
            backup_path = os.path.join(self.backup_dir, backup_filename)
            
            self.update_progress(30, "جاري نسخ الملف...")
            shutil.copy2(DB_NAME, backup_path)
            
            file_size = os.path.getsize(backup_path)
            
            self.update_progress(70, "جاري حفظ المعلومات...")
            
            self.db.execute_insert(
                """INSERT INTO backups 
                   (file_name, backup_type, user_id, file_size, file_path, status) 
                   VALUES (?, 'local', ?, ?, ?, 'completed')""",
                (backup_filename, self.current_user['id'], file_size, backup_path)
            )
            
            self.update_progress(100, "✅ تم إنشاء النسخة الاحتياطية بنجاح", COLORS['success'])
            
            self.db.log_action(self.current_user['id'], 'backup_local',
                              f'إنشاء نسخة احتياطية محلية {backup_filename}')
            
            self.show_snack_bar("تم إنشاء النسخة الاحتياطية المحلية بنجاح", COLORS['success'])
            self.load_backups()
            
            time.sleep(3)
            self.hide_progress()
            
        except Exception as ex:
            self.update_progress(0, f"❌ فشل: {str(ex)}", COLORS['danger'])
            self.show_snack_bar(f"فشل إنشاء النسخة الاحتياطية: {str(ex)}", COLORS['danger'])
    
    def create_cloud_backup(self, e):
        """إنشاء نسخة احتياطية سحابية"""
        if not MEGA_AVAILABLE:
            self.show_snack_bar("مكتبة MEGA غير مثبتة", COLORS['danger'])
            return
        
        mega_email = self.db.get_app_setting('mega_email', MEGA_EMAIL)
        mega_password = self.db.get_app_setting('mega_password', MEGA_PASSWORD)
        
        if not mega_email or not mega_password:
            self.show_snack_bar("بيانات MEGA غير موجودة. أضفها في ملف .env", COLORS['danger'])
            return
        
        # إظهار شريط التقدم
        self.backup_progress.value = 0
        self.backup_status.value = "جاري الاتصال بـ MEGA..."
        self.backup_status.color = COLORS['primary']
        self.page.update()
        
        # تنفيذ في thread منفصل
        def backup_thread():
            try:
                mega = Mega()
                m = mega.login(mega_email, mega_password)
                self.update_progress(20, "جاري إنشاء النسخة المحلية...")
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"backup_cloud_{timestamp}.db"
                backup_path = os.path.join(self.backup_dir, backup_filename)
                
                shutil.copy2(DB_NAME, backup_path)
                file_size = os.path.getsize(backup_path)
                self.update_progress(50, "جاري الرفع إلى MEGA...")
                
                file = m.upload(backup_path)
                link = m.get_upload_link(file)
                self.update_progress(80, "جاري حفظ المعلومات...")
                
                self.db.execute_insert(
                    """INSERT INTO backups 
                       (file_name, backup_type, user_id, file_size, file_path, mega_link, status) 
                       VALUES (?, 'cloud', ?, ?, ?, ?, 'completed')""",
                    (backup_filename, self.current_user['id'], file_size, backup_path, link)
                )
                
                self.update_progress(100, "✅ تم الرفع إلى MEGA بنجاح", COLORS['success'])
                
                self.db.log_action(self.current_user['id'], 'backup_cloud',
                                  f'إنشاء نسخة احتياطية سحابية {backup_filename}')
                
                self.page.snack_bar = ft.SnackBar(
                    content=ft.Text(f"✅ تم إنشاء النسخة الاحتياطية السحابية بنجاح", color=COLORS['white']),
                    bgcolor=COLORS['success']
                )
                self.page.snack_bar.open = True
                self.page.update()
                
                time.sleep(3)
                self.hide_progress()
                self.load_backups()
                
            except Exception as ex:
                error_message = str(ex)
                if "Invalid email or password" in error_message:
                    msg = "❌ فشل تسجيل الدخول: البريد أو كلمة المرور غير صحيحين"
                elif "timeout" in error_message.lower():
                    msg = "❌ فشل الاتصال: تحقق من اتصالك بالإنترنت"
                elif "disk quota" in error_message.lower():
                    msg = "❌ مساحة التخزين السحابية غير كافية"
                else:
                    msg = f"❌ فشل: {error_message[:100]}"
                
                self.update_progress(0, msg, COLORS['danger'])
                self.page.snack_bar = ft.SnackBar(
                    content=ft.Text(msg, color=COLORS['white']),
                    bgcolor=COLORS['danger']
                )
                self.page.snack_bar.open = True
                self.page.update()
                time.sleep(5)
                self.hide_progress()
        
        threading.Thread(target=backup_thread, daemon=True).start()
    
    def update_progress(self, value, status_text, color=COLORS['primary']):
        """تحديث شريط التقدم"""
        self.backup_progress.value = value
        self.backup_status.value = status_text
        self.backup_status.color = color
        self.page.update()
    
    def hide_progress(self):
        """إخفاء شريط التقدم"""
        if self.backup_progress:
            self.backup_progress.value = 0
        if self.backup_status:
            self.backup_status.value = ""
        self.page.update()
    
    def load_backups(self):
        """تحميل سجل النسخ الاحتياطي"""
        if not self.backup_tree:
            return
        
        self.backup_tree.rows.clear()
        
        backups = self.db.execute_query("""
            SELECT b.created_at, b.file_name, b.backup_type, b.file_size, 
                   b.mega_link, b.status, u.username
            FROM backups b
            LEFT JOIN users u ON b.user_id = u.id
            ORDER BY b.created_at DESC 
            LIMIT 50
        """)
        
        for backup in backups:
            created_at, filename, btype, file_size, mega_link, status, username = backup
            
            type_text = "محلي" if btype == 'local' else "سحابي"
            status_text = "✓ مكتمل" if status == 'completed' else "✗ فشل"
            
            if file_size:
                if file_size < 1024:
                    size_text = f"{file_size} B"
                elif file_size < 1024 * 1024:
                    size_text = f"{file_size / 1024:.1f} KB"
                else:
                    size_text = f"{file_size / (1024*1024):.1f} MB"
            else:
                size_text = "-"
            
            link_text = mega_link[:30] + "..." if mega_link and len(mega_link) > 30 else (mega_link or "-")
            
            self.backup_tree.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(created_at[:19] if created_at else "", size=12)),
                        ft.DataCell(ft.Text(filename, size=12)),
                        ft.DataCell(ft.Text(type_text, size=12)),
                        ft.DataCell(ft.Text(size_text, size=12)),
                        ft.DataCell(ft.Text(link_text, size=12)),
                        ft.DataCell(ft.Container(
                            content=ft.Text(status_text, size=12, color=COLORS['white']),
                            bgcolor=COLORS['success'] if status == 'completed' else COLORS['danger'],
                            padding=ft.padding.symmetric(horizontal=8, vertical=2),
                            border_radius=4
                        )),
                        ft.DataCell(ft.Text(username or "", size=12)),
                    ]
                )
            )
        
        self.page.update()
    
    # ================================ دوال مساعدة للنوافذ ================================
    def close_dialog(self, dialog):
        """إغلاق نافذة الحوار"""
        dialog.open = False
        self.page.update()

# ================================ إنشاء ملف .env ================================
def create_env_file():
    """إنشاء ملف .env إذا لم يكن موجوداً"""
    env_path = Path('.env')
    if not env_path.exists():
        env_content = """# إعدادات MEGA للنسخ الاحتياطي السحابي
MEGA_EMAIL=your_email@example.com
MEGA_PASSWORD=your_password

# إعدادات التطبيق
APP_NAME=نظام إدارة العربات اليدوية - الحرم المكي الشريف
COMPANY_NAME=الرئاسة العامة لشؤون المسجد الحرام والمسجد النبوي
"""
        env_path.write_text(env_content, encoding='utf-8')
        print("✅ تم إنشاء ملف .env - يرجى تحديث بيانات MEGA فيه")

# ================================ نقطة البداية ================================
def main(page: ft.Page):
    # إنشاء ملف .env إذا لم يكن موجوداً
    create_env_file()
    
    # تشغيل التطبيق
    app = CartsManagementApp(page)

if __name__ == "__main__":
    ft.app(target=main)
